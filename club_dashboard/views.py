from django.shortcuts import render, redirect, get_object_or_404
from accounts.models import UserProfile, StudentProfile, CoachProfile, ReceptionistProfile, AdministrativeProfile, \
    AccountantProfile
from django.contrib.auth.models import User
from .forms import StudentProfileForm, ArticleModelForm, ServicesModelForm, ServicesClassificationModelForm, \
    ProductsModelForm, ProductsClassificationModelForm, ReceptionistProfileForm, ProductShipmentForm, \
    AdministratorProfileForm, AccountantProfileForm
from accounts.forms import ReceptionistSignupForm, AdministratorSignupForm
from students.models import Blog, ServicesModel, ServicesClassificationModel, ProductsModel, \
    ProductsClassificationModel, ProductsImage, ServicesImage, Order, OrderItem, ServiceOrderModel, OrderCancellation
from django.utils import timezone
# Create your views here.
from django.contrib import messages  # ✅ Fix missing import
from .forms import DirectorProfileForm  # ✅ Fix missing import
from accounts.models import DirectorProfile  # ✅ Add this import
from .models import Notification  # Import the Notification model
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from accounts.models import UserProfile
from club_dashboard.models import Notification  # ✅ Import Notification Model
from .utils import send_notification  # ✅ Import notification function
from django.contrib.auth.decorators import login_required  # ✅ Fix missing import
from django.db.models import Avg
from club_dashboard.models import Review  # ✅ Import Review model from students app
from .models import SalonAppointment, ProductShipment, DashboardSettings
from django.shortcuts import render
from django.db.models import Sum, F, FloatField, Case, When, IntegerField, Value
from django.db.models.functions import Cast
from .models import ProductsModel, ProductImg
from django.utils import timezone
from django.contrib import messages
from .forms import ProductsModelForm
from django.core.paginator import Paginator
import base64
import time
import decimal
from django.core.files.base import ContentFile
from django.utils import timezone
from .forms import ServicesModelForm
from datetime import datetime, timedelta
from django.db import models, transaction
from receptionist_dashboard.models import BookingService
from django.template.loader import render_to_string
import json
from django.http import HttpResponseForbidden
from django.urls import reverse
from accounts.models import UserProfile, DirectorProfile
from .forms import DirectorProfileForm
from django.utils import translation
from receptionist_dashboard.models import SalonBooking
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.utils.translation import gettext as _
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import DetailView
from django.db.models import Count
from accounts.models import ClubsModel
from .decorators import club_permission_required


# Helper function to get user's club
def get_user_club(user):
    user_profile = user.userprofile
    club = None
    if user_profile.account_type == '3':  # Student
        club = user_profile.student_profile.club if hasattr(user_profile, 'student_profile') else None
    elif user_profile.account_type == '4':  # Coach
        club = user_profile.Coach_profile.club if hasattr(user_profile, 'Coach_profile') else None
    elif user_profile.account_type == '2':  # Director
        club = user_profile.director_profile.club if hasattr(user_profile, 'director_profile') else None
    elif user_profile.account_type == '5':  # Receptionist
        club = user_profile.receptionist_profile.club if hasattr(user_profile, 'receptionist_profile') else None
    return club


from coach_dashboard.models import CoachReceptionistTicket
from django.db.models import ExpressionWrapper
from django.db.models import F, ExpressionWrapper, Avg, DurationField
from django.db.models.functions import TruncDay


@club_permission_required('club_dashboard_index')
@login_required
def club_dashboard_index(request):
    context = {}
    user = request.user

    # ✅ Ensure the user has a valid director profile

    # ✅ Get the correct club for the director
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None) or getattr(user.userprofile.custom_role_profile, 'club',
                                                                          None)
    club_name = club.name

    club_admin = user.userprofile.director_profile

    # Get date range (last 30 days by default)
    end_date = timezone.now()
    start_date = end_date - timezone.timedelta(days=30)

    # Revenue data for charts
    daily_revenue = (
        Order.objects.filter(
            club=club,
            status__in=['confirmed', 'completed'],
            created_at__range=[start_date, end_date]
        )
        .annotate(day=TruncDay('created_at'))
        .values('day')
        .annotate(total=Sum('total_price'))
        .order_by('day')
    )

    # Prepare data for revenue chart
    revenue_dates = [entry['day'].strftime('%Y-%m-%d') for entry in daily_revenue]
    revenue_amounts = [float(entry['total']) for entry in daily_revenue]

    # Order status distribution
    status_counts = (
        Order.objects.filter(club=club, created_at__range=[start_date, end_date])
        .values('status')
        .annotate(count=Count('id'))
    )

    status_labels = []
    status_data = []
    status_colors = []
    for status in status_counts:
        status_labels.append(dict(Order.STATUS_CHOICES).get(status['status'], status['status']))
        status_data.append(status['count'])
        if status['status'] == 'completed':
            status_colors.append('#10B981')  # green
        elif status['status'] == 'confirmed':
            status_colors.append('#3B82F6')  # blue
        elif status['status'] == 'pending':
            status_colors.append('#F59E0B')  # yellow
        else:
            status_colors.append('#6B7280')  # gray

    # Ticket status distribution
    ticket_status_counts = (
        CoachReceptionistTicket.objects.filter(
            coach__club=club,
            created_at__range=[start_date, end_date],
        )
        .values('status')
        .annotate(count=Count('id'))
    )

    ticket_labels = []
    ticket_data = []
    ticket_colors = []
    for status in ticket_status_counts:
        ticket_labels.append(status['status'].capitalize())
        ticket_data.append(status['count'])
        if status['status'] == 'resolved':
            ticket_colors.append('#10B981')  # green
        elif status['status'] == 'active':
            ticket_colors.append('#3B82F6')  # blue
        elif status['status'] == 'pending':
            ticket_colors.append('#F59E0B')  # yellow
        else:
            ticket_colors.append('#6B7280')  # gray

    # Revenue by payment method
    payment_method_counts = (
        Order.objects.filter(
            club=club,
            status__in=['confirmed', 'completed'],
            created_at__range=[start_date, end_date]
        )
        .values('payment_method')
        .annotate(total=Sum('total_price'))
    )

    payment_labels = []
    payment_data = []
    payment_colors = ['#3B82F6', '#10B981', '#F59E0B', '#6366F1']

    for method in payment_method_counts:
        payment_labels.append(dict(Order.PAYMENT_METHOD_CHOICES).get(
            method['payment_method'], method['payment_method']
        ))
        payment_data.append(float(method['total']))

    # Weekly revenue comparison
    current_week_start = end_date - timezone.timedelta(days=7)
    previous_week_start = current_week_start - timezone.timedelta(days=7)

    current_week_revenue = (
            Order.objects.filter(
                club=club,
                status__in=['confirmed', 'completed'],
                created_at__range=[current_week_start, end_date]
            )
            .aggregate(total=Sum('total_price'))['total'] or 0
    )

    previous_week_revenue = (
            Order.objects.filter(
                club=club,
                status__in=['confirmed', 'completed'],
                created_at__range=[previous_week_start, current_week_start]
            )
            .aggregate(total=Sum('total_price'))['total'] or 0
    )

    weekly_change = 0
    if previous_week_revenue > 0:
        weekly_change = ((current_week_revenue - previous_week_revenue) / previous_week_revenue) * 100

    current_month_start = end_date.replace(day=1)
    previous_month_start = (current_month_start - timezone.timedelta(days=1)).replace(day=1)
    previous_month_end = current_month_start - timezone.timedelta(days=1)

    current_month_revenue = (
            Order.objects.filter(
                club=club,
                status__in=['confirmed', 'completed'],
                created_at__range=[current_month_start, end_date]
            )
            .aggregate(total=Sum('total_price'))['total'] or 0
    )

    previous_month_revenue = (
            Order.objects.filter(
                club=club,
                status__in=['confirmed', 'completed'],
                created_at__range=[previous_month_start, previous_month_end]
            )
            .aggregate(total=Sum('total_price'))['total'] or 0
    )

    monthly_change = 0
    if previous_month_revenue > 0:
        monthly_change = ((current_month_revenue - previous_month_revenue) / previous_month_revenue) * 100

    # Monthly revenue comparison
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month_start = (current_month_start + timezone.timedelta(days=32)).replace(day=1)
    current_month_end = next_month_start - timezone.timedelta(days=1)

    previous_month_start = (current_month_start - timezone.timedelta(days=1)).replace(day=1)
    previous_month_end = current_month_start - timezone.timedelta(days=1)

    current_month_revenue = (
            Order.objects.filter(
                club=club,
                status__in=['confirmed', 'completed'],
                created_at__range=[current_month_start, current_month_end]
            )
            .aggregate(total=Sum('total_price'))['total'] or 0
    )

    previous_month_revenue = (
            Order.objects.filter(
                club=club,
                status__in=['confirmed', 'completed'],
                created_at__range=[previous_month_start, previous_month_end]
            )
            .aggregate(total=Sum('total_price'))['total'] or 0
    )

    monthly_change = 0
    if previous_month_revenue > 0:
        monthly_change = ((current_month_revenue - previous_month_revenue) / previous_month_revenue) * 100

    # ✅ Get directors linked through UserProfile
    directors = UserProfile.objects.filter(account_type='6', administrator_profile__club=club)
    director_count = directors.count()

    # ✅ Fetch students and coaches from this club
    students = UserProfile.objects.filter(
        account_type='3',
        student_profile__club=club
    ).select_related('user', 'student_profile')
    student_count = students.count()

    coaches = CoachProfile.objects.filter(
        club=club,
        approval_status__in=['pending', 'approved']
    )
    coach_count = coaches.count()

    # ✅ FIXED: Get notifications WITHOUT marking them as read immediately
    notifications = Notification.objects.filter(club=club).order_by('-created_at')
    unread_count = notifications.filter(is_read=False).count()

    # ✅ Don't mark as read here - only mark as read when user actually opens the dropdown
    # notifications.update(is_read=True)  # <-- REMOVE THIS LINE

    def calc_percent(v, total):
        return round((v / total) * 100, 2) if total > 0 else 0

    top_rated_coaches = (
        CoachProfile.objects.filter(club=club)
        .annotate(avg_rating=Avg('coach_reviews__rating'))
        .filter(avg_rating__isnull=False)
        .order_by('-avg_rating')[:5]
    )

    top_reviews = (
        Review.objects.filter(coach__club=club)
        .order_by('-rating', '-created_at')[:5]
    )

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.now().strftime('%Y-%m-%d')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
    except ValueError:
        start_date_obj = timezone.now() - timezone.timedelta(days=30)
        end_date_obj = timezone.now()

    orders = Order.objects.filter(
        club=club,
        created_at__gte=start_date_obj,
        created_at__lte=end_date_obj
    ).order_by('-created_at')

    # Calculate financial metrics
    confirmed_orders = orders.filter(status__in=['confirmed', 'completed'])
    total_revenue = confirmed_orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
    total_commission = confirmed_orders.aggregate(Sum('total_vendor_commission'))['total_vendor_commission__sum'] or 0
    net_revenue = total_revenue - total_commission

    # Calculate average daily sales
    days_in_period = (end_date_obj - start_date_obj).days or 30
    average_daily_sales = round(total_revenue / days_in_period, 2) if days_in_period > 0 else 0

    # Calculate conversion rate (example: orders vs visitors)
    # You'll need to implement actual visitor tracking for this
    total_visitors = 1000  # Replace with actual visitor count if available
    conversion_rate = round((confirmed_orders.count() / total_visitors * 100), 1) if total_visitors > 0 else 0

    # Revenue sources breakdown (example values - customize based on your data)
    revenue_sources = {
        'direct': confirmed_orders.filter(payment_method='credit_card').count(),
        'organic': confirmed_orders.filter(payment_method='cash_on_delivery').count(),
        'referral': 62,  # Replace with actual referral count if available
        'campaign': 1.2  # Replace with actual campaign percentage if available
    }

    # Ticket metrics
    last_7_days = timezone.now() - timezone.timedelta(days=7)
    total_tickets = CoachReceptionistTicket.objects.filter(
        created_at__gte=last_7_days,
        coach__club=club
    ).count()

    new_tickets = CoachReceptionistTicket.objects.filter(
        created_at__gte=last_7_days,
        coach__club=club,
        status='pending'
    ).count()

    open_tickets = CoachReceptionistTicket.objects.filter(
        coach__club=club,
        status__in=['active', 'pending']
    ).count()

    # Calculate average response time (in days)
    resolved_tickets = CoachReceptionistTicket.objects.filter(
        coach__club=club,
        status='resolved'
    )
    if resolved_tickets.exists():
        avg_response_hours = resolved_tickets.annotate(
            response_time=ExpressionWrapper(
                F('resolution_time') - F('created_at'),
                output_field=DurationField()
            )
        ).aggregate(avg_response=Avg('response_time'))['avg_response']

        response_time = round(avg_response_hours.total_seconds() / 3600 / 24, 1)  # Convert to days
    else:
        response_time = 0

    # Monthly campaign metrics (example values)
    social_visitors_percentage = 8.52
    total_emails = 12346
    email_growth_rate = 0.3

    if club:
        directors = UserProfile.objects.filter(
            account_type='2',
            director_profile__club=club
        ).select_related('user', 'director_profile')

        receptionists = UserProfile.objects.filter(
            account_type='5',
            receptionist_profile__club=club
        ).select_related('user', 'receptionist_profile')

        administrators = UserProfile.objects.filter(
            account_type='6',
            administrator_profile__club=club
        ).select_related('user', 'administrator_profile')

        accountants = UserProfile.objects.filter(
            account_type='7',
            accountant_profile__club=club
        ).select_related('user', 'accountant_profile')

        vendorManagers = UserProfile.objects.filter(
            account_type='8',
            vendor_manager_profile__club=club
        ).select_related('user', 'vendor_manager_profile')

        customRoles = UserProfile.objects.filter(
            account_type='9',
            custom_role_profile__club=club
        ).select_related('user', 'custom_role_profile')

        staff_list = []

        for director in directors:
            staff_list.append({
                'userprofile': director,
                'role': 'مدير عام',
                'role_en': 'General Manager',
                'profile': director.director_profile,
                'profile_type': 'director'
            })

        for receptionist in receptionists:
            staff_list.append({
                'userprofile': receptionist,
                'role': 'موظف استقبال',
                'role_en': 'Receptionist',
                'profile': receptionist.receptionist_profile,
                'profile_type': 'receptionist'
            })

        for administrator in administrators:
            staff_list.append({
                'userprofile': administrator,
                'role': 'إداري',
                'role_en': 'Administrator',
                'profile': administrator.administrator_profile,
                'profile_type': 'administrator'
            })

        for accountant in accountants:
            staff_list.append({
                'userprofile': accountant,
                'role': 'محاسب',
                'role_en': 'Accountant',
                'profile': accountant.accountant_profile,
                'profile_type': 'accountant'
            })

        for vendorManager in vendorManagers:
            staff_list.append({
                'userprofile': vendorManager,
                'role': 'مدير تجار',
                'role_en': 'Vendor Manager',
                'profile': vendorManager.vendor_manager_profile,
                'profile_type': 'vendorManager'
            })

        for customRole in customRoles:
            staff_list.append({
                'userprofile': customRole,
                'role': 'دور خاص',
                'role_en': 'Custom Role',
                'profile': customRole.custom_role_profile,
                'profile_type': 'customRole'
            })

        staff_list.sort(key=lambda x: x['userprofile'].creation_date, reverse=True)
    else:
        staff_list = []

    categories_count = Category.objects.filter(is_active=True).count()
    subcategories_count = SubCategory.objects.filter(is_active=True).count()

    total_revenue = int(
        orders.filter(status__in=['confirmed', 'completed']).aggregate(Sum('total_price'))['total_price__sum'] or 0)
    context['LANGUAGE_CODE'] = translation.get_language()

    return render(request, 'club_dashboard/index.html', {
        'staff_list': staff_list,
        'clubName': club_name,
        'club': club,
        'students': students,
        'coaches': coaches,
        'directors': directors,
        'director_count': directors.count(),
        'coach_count': coach_count,
        'student_count': student_count,
        'orders': orders,
        'notifications': notifications,
        'unread_count': unread_count,
        'top_rated_coaches': top_rated_coaches,
        'top_reviews': top_reviews,
        'total_revenue': total_revenue,
        'total_commission': total_commission,
        'net_revenue': net_revenue,
        'average_daily_sales': average_daily_sales,
        'conversion_rate': conversion_rate,
        'revenue_sources': revenue_sources,
        'total_tickets': total_tickets,
        'new_tickets': new_tickets,
        'categories_count': categories_count,
        'subcategories_count': subcategories_count,
        'open_tickets': open_tickets,
        'response_time': response_time,
        'social_visitors_percentage': social_visitors_percentage,
        'total_emails': total_emails,
        'email_growth_rate': email_growth_rate,
        'LANGUAGE_CODE': translation.get_language(),
        'club_admin': club_admin,
        # New chart data
        'revenue_dates': revenue_dates,
        'revenue_amounts': revenue_amounts,
        'status_labels': status_labels,
        'status_data': status_data,
        'status_colors': status_colors,
        'ticket_labels': ticket_labels,
        'ticket_data': ticket_data,
        'ticket_colors': ticket_colors,
        'payment_labels': payment_labels,
        'payment_data': payment_data,
        'payment_colors': payment_colors,
        'current_week_revenue': current_week_revenue,
        'previous_week_revenue': previous_week_revenue,
        'weekly_change': weekly_change,
        'current_month_revenue': current_month_revenue,
        'previous_month_revenue': previous_month_revenue,
        'monthly_change': monthly_change,
        'current_month_revenue': current_month_revenue,
        'previous_month_revenue': previous_month_revenue,
        'monthly_change': monthly_change,
    })


from .models import LandingPageContent, LandingPageFeature, LandingPageBanner, LandingPageFAQ, NavItem
from .forms import (LandingPageContentForm, LandingPageFeatureForm,
                    LandingPageBannerForm, LandingPageFAQForm, NavItemForm)


@login_required
def edit_landing_content(request):
    """Director dashboard view to edit landing page content"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    # Get or create landing content for this club
    landing_content, created = LandingPageContent.objects.get_or_create()

    if request.method == 'POST':
        form = LandingPageContentForm(request.POST, request.FILES, instance=landing_content)
        if form.is_valid():
            form.save()
            messages.success(request, 'Landing page content updated successfully!')
            return redirect('edit_landing_content')
    else:
        form = LandingPageContentForm(instance=landing_content)

    # Get related items
    features = landing_content.features.all()
    banners = landing_content.banners.all()
    faqs = landing_content.faqs.all()
    nav_items = landing_content.nav_items.all()

    context = {
        'form': form,
        'landing_content': landing_content,
        'features': features,
        'banners': banners,
        'faqs': faqs,
        'nav_items': nav_items,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/edit_landing_content.html', context)


@login_required
def manage_features(request):
    """Manage landing page features"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    landing_content = LandingPageContent.objects.get_or_create()[0]

    if request.method == 'POST':
        form = LandingPageFeatureForm(request.POST)
        if form.is_valid():
            feature = form.save(commit=False)
            feature.landing_content = landing_content
            feature.save()
            messages.success(request, 'Feature added successfully!')
            return redirect('manage_features')
    else:
        form = LandingPageFeatureForm()

    features = landing_content.features.all()

    context = {
        'form': form,
        'features': features,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/manage_features.html', context)


@login_required
def edit_feature(request, feature_id):
    """Edit a specific feature"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    feature = get_object_or_404(LandingPageFeature, id=feature_id)

    if request.method == 'POST':
        form = LandingPageFeatureForm(request.POST, instance=feature)
        if form.is_valid():
            form.save()
            messages.success(request, 'Feature updated successfully!')
            return redirect('manage_features')
    else:
        form = LandingPageFeatureForm(instance=feature)

    context = {
        'form': form,
        'feature': feature,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/edit_feature.html', context)


@login_required
def delete_feature(request, feature_id):
    """Delete a feature"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    feature = get_object_or_404(LandingPageFeature, id=feature_id)
    feature.delete()
    messages.success(request, 'Feature deleted successfully!')
    return redirect('manage_features')


@login_required
def manage_banners(request):
    """Manage landing page banners"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    landing_content = LandingPageContent.objects.get_or_create()[0]

    if request.method == 'POST':
        form = LandingPageBannerForm(request.POST, request.FILES)
        if form.is_valid():
            banner = form.save(commit=False)
            banner.landing_content = landing_content
            banner.save()
            messages.success(request, 'Banner added successfully!')
            return redirect('manage_banners')
    else:
        form = LandingPageBannerForm()

    banners = landing_content.banners.all()

    context = {
        'form': form,
        'banners': banners,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/manage_banners.html', context)


@login_required
def delete_banner(request, banner_id):
    """Delete a banner"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    banner = get_object_or_404(LandingPageBanner, id=banner_id)
    banner.delete()
    messages.success(request, 'Banner deleted successfully!')
    return redirect('manage_banners')


@login_required
def manage_faqs(request):
    """Manage landing page FAQs"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    landing_content = LandingPageContent.objects.get_or_create()[0]

    if request.method == 'POST':
        form = LandingPageFAQForm(request.POST)
        if form.is_valid():
            faq = form.save(commit=False)
            faq.landing_content = landing_content
            faq.save()
            messages.success(request, 'FAQ added successfully!')
            return redirect('manage_faqs')
    else:
        form = LandingPageFAQForm()

    faqs = landing_content.faqs.all()

    context = {
        'form': form,
        'faqs': faqs,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/manage_faqs.html', context)


@login_required
def delete_faq(request, faq_id):
    """Delete a FAQ"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    faq = get_object_or_404(LandingPageFAQ, id=faq_id)
    faq.delete()
    messages.success(request, 'FAQ deleted successfully!')
    return redirect('manage_faqs')


@login_required
def manage_nav_items(request):
    """Manage navigation items"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    landing_content = LandingPageContent.objects.get_or_create()[0]

    if request.method == 'POST':
        form = NavItemForm(request.POST)
        if form.is_valid():
            nav_item = form.save(commit=False)
            nav_item.landing_content = landing_content
            nav_item.save()
            messages.success(request, 'Navigation item added successfully!')
            return redirect('manage_nav_items')
    else:
        form = NavItemForm()

    nav_items = landing_content.nav_items.all()

    context = {
        'form': form,
        'nav_items': nav_items,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/manage_nav_items.html', context)


@login_required
def delete_nav_item(request, nav_item_id):
    """Delete a navigation item"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None)

    if not club:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('club_dashboard_index')

    nav_item = get_object_or_404(NavItem, id=nav_item_id)
    nav_item.delete()
    messages.success(request, 'Navigation item deleted successfully!')
    return redirect('manage_nav_items')


from .models import ClubContact
from .forms import ClubContactForm


@login_required
def update_club_contact(request, club_id):
    """Update club contact information"""
    print(f"DEBUG: update_club_contact called with club_id={club_id}")
    print(f"DEBUG: User authenticated: {request.user.is_authenticated}")
    print(f"DEBUG: User ID: {request.user.id}")
    print(f"DEBUG: User type: {type(request.user)}")

    user = request.user
    print(f"DEBUG: User object: {user}")

    try:
        club = get_object_or_404(ClubsModel, id=club_id)
        print(f"DEBUG: Club found: {club.id} - {club.name}")
    except Exception as e:
        print(f"DEBUG ERROR: Failed to get club: {e}")
        raise

    # Check if user has permission to edit this club
    print(f"DEBUG: Checking user permissions...")
    print(f"DEBUG: User has userprofile: {hasattr(user, 'userprofile')}")

    if hasattr(user, 'userprofile'):
        print(f"DEBUG: UserProfile exists: {user.userprofile}")
        print(f"DEBUG: UserProfile type: {type(user.userprofile)}")
        print(f"DEBUG: UserProfile has director_profile: {hasattr(user.userprofile, 'director_profile')}")

        if hasattr(user.userprofile, 'director_profile'):
            director_profile = user.userprofile.director_profile
            print(f"DEBUG: Director profile found: {director_profile}")
            print(f"DEBUG: Director's club: {getattr(director_profile, 'club', 'NO CLUB ATTRIBUTE')}")
            print(f"DEBUG: Expected club_id: {club_id}")
            print(f"DEBUG: Director's club ID: {getattr(getattr(director_profile, 'club', None), 'id', 'NO CLUB ID')}")

    if not hasattr(user, 'userprofile') or not hasattr(user.userprofile,
                                                       'director_profile') or user.userprofile.director_profile.club != club:
        print(f"DEBUG: Permission denied for user {user.id} to edit club {club_id}")
        print(f"DEBUG: Reason check:")
        print(f"  - Has userprofile: {hasattr(user, 'userprofile')}")
        if hasattr(user, 'userprofile'):
            print(f"  - Has director_profile: {hasattr(user.userprofile, 'director_profile')}")
            if hasattr(user.userprofile, 'director_profile'):
                print(f"  - Club match: {user.userprofile.director_profile.club == club}")
        messages.error(request, 'You do not have permission to edit this club.')
        return redirect('club_dashboard_index')

    print(f"DEBUG: User has permission to edit club {club_id}")

    # Get or create contact info
    print(f"DEBUG: Getting or creating ClubContact...")
    try:
        contact_info, created = ClubContact.objects.get_or_create()
        print(f"DEBUG: Contact info {'created' if created else 'retrieved'}: {contact_info.id}")
    except Exception as e:
        print(f"DEBUG ERROR: Failed to get/create ClubContact: {e}")
        messages.error(request, 'Error accessing contact information.')
        return redirect('ViewClubProfile', id=club.id)

    if request.method == 'POST':
        print(f"DEBUG: POST request received")
        print(f"DEBUG: POST data: {request.POST}")
        print(f"DEBUG: FILES data: {request.FILES}")

        form = ClubContactForm(request.POST, request.FILES, instance=contact_info)
        print(f"DEBUG: Form created with instance: {form.instance}")

        if form.is_valid():
            print(f"DEBUG: Form is valid")
            try:
                saved_instance = form.save()
                print(f"DEBUG: Form saved successfully. Instance ID: {saved_instance.id}")
                messages.success(request, 'Contact information updated successfully!')
                return redirect('ViewClubProfile', id=club.id)
            except Exception as e:
                print(f"DEBUG ERROR: Failed to save form: {e}")
                messages.error(request, 'Error saving contact information.')
        else:
            print(f"DEBUG: Form is INVALID")
            print(f"DEBUG: Form errors: {form.errors}")
            print(f"DEBUG: Form non-field errors: {form.non_field_errors()}")
            messages.error(request, 'Please correct the errors below.')
    else:
        print(f"DEBUG: GET request received")
        form = ClubContactForm(instance=contact_info)
        print(f"DEBUG: Form created for GET with instance ID: {contact_info.id}")

    context = {
        'form': form,
        'club': club,
        'contact_info': contact_info,
        'LANGUAGE_CODE': translation.get_language(),
    }

    print(f"DEBUG: Context prepared with club_id={club.id}")
    print(f"DEBUG: LANGUAGE_CODE: {translation.get_language()}")

    return render(request, 'accounts/profiles/Club/ViewClubProfile.html', context)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json


@login_required
@require_POST
def mark_notifications_read(request):
    """Mark all notifications as read for the current user's club."""
    try:
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(
            user.userprofile.administrator_profile, 'club', None) or getattr(user.userprofile.vendor_manager_profile,
                                                                             'club', None)

        if not club:
            return JsonResponse({'success': False, 'error': 'No club found'})

        # Mark all unread notifications as read
        updated_count = Notification.objects.filter(
            club=club,
            is_read=False
        ).update(is_read=True)

        return JsonResponse({
            'success': True,
            'updated_count': updated_count
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@club_permission_required('delete_review')
@login_required
def delete_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)

    user = request.user
    if hasattr(user, 'userprofile') and user.userprofile.director_profile:
        if review.coach.club == user.userprofile.director_profile.club:
            review.delete()
            messages.success(request, "Review deleted successfully.")
        else:
            messages.error(request, "You are not authorized to delete this review.")
    else:
        messages.error(request, "Unauthorized action.")

    return redirect('club_dashboard_index')  # Use your correct dashboard URL name


@club_permission_required('viewStudents')
def viewStudents(request):
    context = {}
    """Displays all students in the club."""
    user = request.user

    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    students = UserProfile.objects.filter(
        account_type='3',
        student_profile__club=club
    ).select_related('user', 'student_profile')

    # ✅ Enrich each student with subscription status
    for student in students:
        profile = student.student_profile
        if profile and hasattr(profile, 'get_subscription_status'):
            student.subscription_status = profile.get_subscription_status()
            student.manual_status_display = profile.get_manual_status_display()
        else:
            student.subscription_status = "unknown"
            student.manual_status_display = "-"
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/viewStudents.html', {'students': students, 'club': club})


import pandas as pd
import openpyxl
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User
from django.db import transaction
from datetime import datetime
import io


# Add this view for handling the import
def import_students(request):
    """Display import students page"""
    context = {}
    user = request.user

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        messages.error(request, "No club associated with your account.")
        return redirect('viewStudents')

    context['club'] = club
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/import_students.html', context)


def process_import_students(request):
    """Process the uploaded Excel/CSV file"""
    if request.method != 'POST':
        return redirect('import_students')

    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        messages.error(request, "No club associated with your account.")
        return redirect('viewStudents')

    if 'file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('import_students')

    file = request.FILES['file']

    # Validate file type
    if not file.name.endswith(('.xlsx', '.xls', '.csv')):
        messages.error(request, "Please upload an Excel (.xlsx, .xls) or CSV file.")
        return redirect('import_students')

    try:
        # Read the file
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        # Validate required columns
        required_columns = ['username', 'email', 'full_name', 'phone', 'birthday']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
            return redirect('import_students')

        # Process the data
        success_count = 0
        error_count = 0
        errors = []

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Check if username already exists
                    if User.objects.filter(username=row['username']).exists():
                        errors.append(f"Row {index + 2}: Username '{row['username']}' already exists")
                        error_count += 1
                        continue

                    # Check if email already exists
                    if User.objects.filter(email=row['email']).exists():
                        errors.append(f"Row {index + 2}: Email '{row['email']}' already exists")
                        error_count += 1
                        continue

                    # Parse birthday
                    try:
                        birthday = pd.to_datetime(row['birthday']).date()
                    except:
                        errors.append(f"Row {index + 2}: Invalid birthday format")
                        error_count += 1
                        continue

                    # Create User
                    user_obj = User.objects.create_user(
                        username=row['username'],
                        email=row['email'],
                        password='defaultpass123'  # You might want to generate random passwords
                    )

                    # Create StudentProfile
                    student_profile = StudentProfile.objects.create(
                        full_name=row['full_name'],
                        phone=str(row['phone']),
                        birthday=birthday,
                        club=club,
                        manual_status=row.get('manual_status', 'trial')
                    )

                    # Create UserProfile
                    UserProfile.objects.create(
                        user=user_obj,
                        account_type='3',  # student
                        student_profile=student_profile,
                        is_active=True
                    )

                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1

        # Show results
        if success_count > 0:
            messages.success(request, f"Successfully imported {success_count} students.")

        if error_count > 0:
            error_message = f"Failed to import {error_count} students:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                error_message += f"\n... and {len(errors) - 10} more errors"
            messages.error(request, error_message)

    except Exception as e:
        messages.error(request, f"Error processing file: {str(e)}")

    return redirect('viewStudents')


def download_sample_template(request):
    """Download a sample Excel template for importing students"""

    # Create sample data
    data = {
        'username': ['client1', 'client2', 'client3'],
        'email': ['client1@example.com', 'client2@example.com', 'client3@example.com'],
        'full_name': ['Ahmed Ali', 'Sara Mohamed', 'Omar Hassan'],
        'phone': ['01234567890', '01098765432', '01156789012'],
        'birthday': ['1995-01-15', '1998-03-22', '2000-07-10'],
        'manual_status': ['trial', 'active', 'trial']
    }

    df = pd.DataFrame(data)

    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Students', index=False)

        # Format the worksheet
        worksheet = writer.sheets['Students']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="clients_import_template.xlsx"'

    return response


def export_students_excel(request):
    user = request.user

    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    students = UserProfile.objects.filter(
        account_type='3',
        student_profile__club=club
    ).select_related('user', 'student_profile')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "العملاء"

    arabic_font = Font(name='Arial', size=12)
    right_align = Alignment(horizontal='right')

    headers = [
        _('Username'), _('Email'), _('Full Name'),
        _('Phone'), _('Birthday'), _('Subscription Status')
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = arabic_font
        cell.alignment = right_align

    for student in students:
        profile = student.student_profile
        subscription = profile.get_subscription_status() if hasattr(profile, 'get_subscription_status') else "unknown"

        row = [
            student.user.username,
            student.user.email,
            profile.full_name if profile else '',
            profile.phone if profile else '',
            str(profile.birthday) if profile and profile.birthday else '',
            subscription
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = arabic_font
            cell.alignment = right_align

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=clients.xlsx'
    wb.save(response)
    return response


from django.db import IntegrityError
from django.core.exceptions import ValidationError


@club_permission_required('addStudent')
# views.py - Update the addStudent view

def addStudent(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    if not club.can_add_more_players:
        messages.error(request,
                       f"لا يمكن إضافة المزيد من العملاء. الحد الأقصى للباقة الحالية هو {club.get_max_players()} عميل.")
        return redirect('viewStudents')

    if request.method == 'POST':
        form = StudentProfileForm(request.POST)
        try:
            if form.is_valid():
                # Check for similar existing students before creating
                full_name = form.cleaned_data['full_name']
                phone = form.cleaned_data['phone']
                email = form.cleaned_data['email']

                # Check for similar students
                similar_students = StudentProfile.objects.filter(
                    models.Q(full_name__iexact=full_name) |
                    models.Q(phone=phone)
                )

                if similar_students.exists():
                    # Create a list of similar students for the warning message
                    similar_list = []
                    for student in similar_students:
                        similar_list.append(f"{student.full_name} - {student.phone}")

                    # Store form data in session to repopulate after warning
                    request.session['pending_student_data'] = {
                        'username': form.cleaned_data['username'],
                        'email': email,
                        'full_name': full_name,
                        'phone': phone,
                        'birthday': form.cleaned_data['birthday'].isoformat() if form.cleaned_data[
                            'birthday'] else None,
                        'gender': form.cleaned_data['gender'],
                        'password': form.cleaned_data.get('password', '')
                    }

                    # Store similar students info in session
                    request.session['similar_students'] = similar_list

                    # Redirect to confirmation page
                    return redirect('confirm_student_duplicate')

                # Create new user if no duplicates found
                username = form.cleaned_data['username']
                password = form.cleaned_data.get('password', None)

                student = User.objects.create(username=username, email=email)
                if password:
                    student.set_password(password)
                student.save()

                # Create and link Student Profile
                student_profile = form.save(commit=False)
                student_profile.user = student
                student_profile.club = club
                student_profile.save()

                # Create UserProfile entry
                UserProfile.objects.create(user=student, account_type='3', student_profile=student_profile)

                messages.success(request, "تم إضافة العميل بنجاح")
                return redirect('viewStudents')

        except IntegrityError as e:
            messages.error(request, "حدث خطأ في قاعدة البيانات. يرجى المحاولة مرة أخرى.")
            logger.error(f"IntegrityError adding student: {str(e)}")
        except ValidationError as e:
            for error in e.error_list:
                messages.error(request, str(error))
        except Exception as e:
            messages.error(request, "حدث خطأ غير متوقع. يرجى المحاولة مرة أخرى.")
            logger.error(f"Unexpected error adding student: {str(e)}")
    else:
        form = StudentProfileForm()
        # Check if we have pending data from a duplicate warning
        if 'pending_student_data' in request.session:
            data = request.session.pop('pending_student_data')
            form = StudentProfileForm(initial=data)

    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/addStudent.html', {
        'form': form,
        'club': club,
    })


# Add this new view for handling duplicate confirmation
def confirm_student_duplicate(request):
    if 'similar_students' not in request.session:
        return redirect('addStudent')

    similar_students = request.session.get('similar_students', [])
    pending_data = request.session.get('pending_student_data', {})

    if request.method == 'POST':
        if 'confirm' in request.POST:
            # User confirmed to proceed despite duplicates
            try:
                # Create new user
                username = pending_data['username']
                email = pending_data['email']
                password = pending_data.get('password', None)

                student = User.objects.create(username=username, email=email)
                if password:
                    student.set_password(password)
                student.save()

                # Create Student Profile
                student_profile = StudentProfile(
                    full_name=pending_data['full_name'],
                    phone=pending_data['phone'],
                    birthday=pending_data['birthday'],
                    gender=pending_data['gender'],
                    user=student,
                    club=request.user.userprofile.director_profile.club
                )
                student_profile.save()

                # Create UserProfile entry
                UserProfile.objects.create(user=student, account_type='3', student_profile=student_profile)

                # Clean up session
                request.session.pop('similar_students', None)
                request.session.pop('pending_student_data', None)

                messages.success(request, "تم إضافة العميل بنجاح على الرغم من وجود تشابه في البيانات")
                return redirect('viewStudents')

            except Exception as e:
                messages.error(request, f"حدث خطأ أثناء إضافة العميل: {str(e)}")
                return redirect('addStudent')

        elif 'cancel' in request.POST:
            # User canceled, redirect back to form with data preserved
            return redirect('addStudent')

    return render(request, 'club_dashboard/students/confirm_duplicate.html', {
        'similar_students': similar_students,
        'student_data': pending_data
    })


@club_permission_required('editStudent')
def editStudent(request, id):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    student_profile = get_object_or_404(StudentProfile, id=id)
    student = get_object_or_404(User, userprofile__student_profile=student_profile)

    if request.method == 'POST':
        form = StudentProfileForm(request.POST, instance=student_profile)
        try:
            if form.is_valid():
                # Update student details
                new_username = form.cleaned_data['username']
                new_email = form.cleaned_data['email']
                password = form.cleaned_data.get('password')

                # Check if username/email changed and if they're unique
                if new_username != student.username and User.objects.filter(username=new_username).exists():
                    raise ValidationError("اسم المستخدم موجود مسبقاً")
                if new_email != student.email and User.objects.filter(email=new_email).exists():
                    raise ValidationError("البريد الإلكتروني مستخدم مسبقاً")

                student.username = new_username
                student.email = new_email
                if password:
                    student.set_password(password)
                student.save()

                student_profile = form.save(commit=False)
                student_profile.user = student
                student_profile.save()

                messages.success(request, "تم تحديث بيانات العميل بنجاح")
                return redirect('viewStudents')

        except IntegrityError as e:
            messages.error(request, "حدث خطأ في قاعدة البيانات. يرجى المحاولة مرة أخرى.")
            logger.error(f"IntegrityError editing student: {str(e)}")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, "حدث خطأ غير متوقع. يرجى المحاولة مرة أخرى.")
            logger.error(f"Unexpected error editing student: {str(e)}")
    else:
        initial_data = {
            'username': student.username,
            'email': student.email,
        }
        form = StudentProfileForm(instance=student_profile, initial=initial_data)

    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/editStudent.html', {
        'form': form,
        'student': student,
        'club': club,
    })


@login_required
@club_permission_required('deleteStudent')
def deleteStudent(request, id):
    """Deletes a student from the club."""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    student_profile = get_object_or_404(StudentProfile, id=id)
    student = get_object_or_404(User, userprofile__student_profile=student_profile)

    student_name = student.username
    student_profile.delete()
    student.delete()
    send_notification(user, club, f" تم حذف العميل 🗑️ {student_name} .من المنصة ")

    messages.success(request, "CLient has been deleted successfully.")
    return redirect('viewStudents')


from django.db.models import Sum, Q, OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce


@club_permission_required('viewCoachs')
def viewCoachs(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Approved coaches
    coach_userprofile = UserProfile.objects.filter(
        account_type='4',
        Coach_profile__club=club
    ).select_related('user', 'Coach_profile').annotate(
        total_sales=Coalesce(
            Subquery(
                OrderItem.objects.filter(
                    Q(order__status='confirmed'),
                    Q(product__creator__userprofile__Coach_profile=OuterRef('Coach_profile')) |
                    Q(service__creator__userprofile__Coach_profile=OuterRef('Coach_profile'))
                ).values('product__creator__userprofile__Coach_profile')
                .annotate(total=Sum('price'))
                .values('total')[:1]
            ),
            0.00,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )

    # Pending coaches (for the top section)
    pending_coaches = CoachProfile.objects.filter(
        club=club,
        approval_status='pending'
    ).order_by('-created_at')[:3]  # Get only 3 most recent pending coaches

    # Statistics
    total_coaches = coach_userprofile.count()
    active_coaches = coach_userprofile.filter(is_active=True).count()
    inactive_coaches = total_coaches - active_coaches
    pending_count = CoachProfile.objects.filter(club=club, approval_status='pending').count()

    context.update({
        'LANGUAGE_CODE': translation.get_language(),
        'coach_userprofile': coach_userprofile,
        'club': club,
        'pending_coaches': pending_coaches,
        'stats': {
            'total': total_coaches,
            'active': active_coaches,
            'inactive': inactive_coaches,
            'pending': pending_count,
        }
    })

    return render(request, 'club_dashboard/coachs/viewCoachs.html', context)


def export_coaches_excel(request):
    user = request.user

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    coaches = UserProfile.objects.filter(
        account_type='4',
        Coach_profile__club=club
    ).select_related('user', 'Coach_profile')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Coaches")

    font = Font(name='Arial', size=12)
    align = Alignment(horizontal='right')

    headers = [
        _('Username'),
        _('Email'),
        _('Full Name'),
        _('Phone'),
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = font
        cell.alignment = align

    for coach in coaches:
        profile = coach.Coach_profile
        row = [
            coach.user.username,
            coach.user.email,
            profile.full_name if profile else '',
            profile.phone if profile else '',
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = font
            cell.alignment = align

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=vendors.xlsx'
    wb.save(response)
    return response


from club_dashboard.forms import CoachProfileForm
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone, translation
from club_dashboard.forms import CoachProfileForm
from accounts.models import UserProfile
import json


@club_permission_required('addCoach')
@login_required
def addCoach(request):
    print("=== addCoach VIEW STARTED ===")
    context = {}
    user = request.user
    print(f"Logged in user: {user}")

    # Get the club associated with the user
    club = (getattr(user.userprofile.director_profile, 'club', None) or
            getattr(user.userprofile.administrator_profile, 'club', None) or
            getattr(user.userprofile.vendor_manager_profile, 'club', None))

    print(f"Club found: {club}")

    if request.method == 'POST':
        print("=== POST request detected ===")
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        print(f"Form data received -> username: {username}, email: {email}, password: {'yes' if password else 'no'}")

        # Check for duplicate username and email
        if User.objects.filter(username=username).exists():
            print("Duplicate username detected!")
            messages.error(request,
                           "اسم المستخدم موجود بالفعل" if translation.get_language() == 'ar' else "Username already exists.")
            context['form'] = CoachProfileForm(request.POST, club=club)
            context['LANGUAGE_CODE'] = translation.get_language()
            context['club'] = club
            return render(request, 'club_dashboard/coachs/addCoach.html', context)

        if User.objects.filter(email=email).exists():
            print("Duplicate email detected!")
            messages.error(request,
                           "البريد الإلكتروني مستخدم بالفعل" if translation.get_language() == 'ar' else "Email is already in use.")
            context['form'] = CoachProfileForm(request.POST, club=club)
            context['LANGUAGE_CODE'] = translation.get_language()
            context['club'] = club
            return render(request, 'club_dashboard/coachs/addCoach.html', context)

        form = CoachProfileForm(request.POST, club=club)
        print("CoachProfileForm initialized")

        if form.is_valid():
            print("Form is valid ✅")

            # Create new user
            coach = User.objects.create(username=username, email=email)
            print(f"New User created: {coach}")

            if password:
                coach.set_password(password)
                print("Password set for new user")

            coach.save()
            print("User saved successfully")

            # Create coach profile
            coach_profile = form.save(commit=False)
            coach_profile.club = club
            coach_profile.approval_status = 'approved'
            coach_profile.approved_by = user
            coach_profile.approved_at = timezone.now()
            print("Coach profile prepared (not yet saved)")

            # Handle branches data
            number_of_branches = form.cleaned_data.get('number_of_branches', 1)
            print(f"Number of branches: {number_of_branches}")

            if number_of_branches > 1:
                branches_json = request.POST.get('branches_data')
                print(f"Branches JSON received: {branches_json}")
                if branches_json:
                    try:
                        branches = json.loads(branches_json)
                        coach_profile.branches = branches
                        print(f"Branches parsed successfully: {branches}")
                    except json.JSONDecodeError:
                        print("JSONDecodeError: Invalid branches data")
                        messages.error(request,
                                       "خطأ في بيانات الفروع" if translation.get_language() == 'ar' else "Error in branches data")
                        context['form'] = form
                        context['LANGUAGE_CODE'] = translation.get_language()
                        context['club'] = club
                        return render(request, 'club_dashboard/coachs/addCoach.html', context)

            coach_profile.save()
            print("Coach profile saved successfully")

            # Create UserProfile entry
            user_profile = UserProfile.objects.create(
                user=coach,
                account_type='4',
                Coach_profile=coach_profile,
                is_active=True
            )
            print(f"UserProfile created: {user_profile}")

            # Send notification
            send_notification(user, club, f"التاجر الجديد 📢 {username} انضم إلى {club.name}.")
            print("Notification sent")

            messages.success(request,
                             "تم إضافة التاجر بنجاح." if translation.get_language() == 'ar' else "Vendor added successfully.")
            print("✅ Coach added successfully — redirecting to viewCoachs")
            return redirect('viewCoachs')
        else:
            print("Form validation failed ❌")
            print(f"Form errors: {form.errors.as_json()}")
            messages.error(request,
                           "يرجى تصحيح الأخطاء في النموذج" if translation.get_language() == 'ar' else "Please correct the errors in the form")
            context['form'] = form
            context['LANGUAGE_CODE'] = translation.get_language()
            context['club'] = club
            return render(request, 'club_dashboard/coachs/addCoach.html', context)
    else:
        print("=== GET request detected ===")
        form = CoachProfileForm(club=club)

    context['LANGUAGE_CODE'] = translation.get_language()
    context['form'] = form
    context['club'] = club
    print("Rendering addCoach.html page")
    print("=== addCoach VIEW ENDED ===")
    return render(request, 'club_dashboard/coachs/addCoach.html', context)


@club_permission_required('editCoach')
@login_required
def editCoach(request, id):
    context = {}
    user = request.user

    # Get the club associated with the user
    club = (getattr(user.userprofile.director_profile, 'club', None) or
            getattr(user.userprofile.administrator_profile, 'club', None) or
            getattr(user.userprofile.vendor_manager_profile, 'club', None))

    try:
        coach_profile = CoachProfile.objects.get(id=id, club=club)
        coach_user = User.objects.get(userprofile__Coach_profile=coach_profile)
    except (CoachProfile.DoesNotExist, User.DoesNotExist):
        messages.error(request, "التاجر غير موجود" if translation.get_language() == 'ar' else "Vendor not found")
        return redirect('viewCoachs')

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        # Check for duplicate username (excluding current user)
        if User.objects.filter(username=username).exclude(id=coach_user.id).exists():
            messages.error(request,
                           "اسم المستخدم موجود بالفعل" if translation.get_language() == 'ar' else "Username already exists.")
            form = CoachProfileForm(request.POST, instance=coach_profile, club=club)
            context.update({
                'form': form,
                'coach': coach_user,
                'coach_profile': coach_profile,
                'LANGUAGE_CODE': translation.get_language(),
                'club': club
            })
            return render(request, 'club_dashboard/coachs/editCoach.html', context)

        # Check for duplicate email (excluding current user)
        if User.objects.filter(email=email).exclude(id=coach_user.id).exists():
            messages.error(request,
                           "البريد الإلكتروني مستخدم بالفعل" if translation.get_language() == 'ar' else "Email is already in use.")
            form = CoachProfileForm(request.POST, instance=coach_profile, club=club)
            context.update({
                'form': form,
                'coach': coach_user,
                'coach_profile': coach_profile,
                'LANGUAGE_CODE': translation.get_language(),
                'club': club
            })
            return render(request, 'club_dashboard/coachs/editCoach.html', context)

        form = CoachProfileForm(request.POST, instance=coach_profile, club=club)

        if form.is_valid():
            # Update user credentials
            coach_user.username = username
            coach_user.email = email
            if password:
                coach_user.set_password(password)
            coach_user.save()

            # Update coach profile
            coach_profile = form.save(commit=False)

            # Handle branches data
            number_of_branches = form.cleaned_data.get('number_of_branches', 1)
            if number_of_branches > 1:
                branches_json = request.POST.get('branches_data')
                if branches_json:
                    try:
                        branches = json.loads(branches_json)
                        coach_profile.branches = branches
                    except json.JSONDecodeError:
                        messages.error(request,
                                       "خطأ في بيانات الفروع" if translation.get_language() == 'ar' else "Error in branches data")
                        context.update({
                            'form': form,
                            'coach': coach_user,
                            'coach_profile': coach_profile,
                            'LANGUAGE_CODE': translation.get_language(),
                            'club': club
                        })
                        return render(request, 'club_dashboard/coachs/editCoach.html', context)
            else:
                # Clear branches if only one branch
                coach_profile.branches = []

            coach_profile.save()

            messages.success(request,
                             "تم تحديث بيانات التاجر بنجاح." if translation.get_language() == 'ar' else "Vendor updated successfully.")
            return redirect('viewCoachs')
        else:
            messages.error(request,
                           "يرجى تصحيح الأخطاء في النموذج" if translation.get_language() == 'ar' else "Please correct the errors in the form")
    else:
        form = CoachProfileForm(instance=coach_profile, club=club)

    context.update({
        'form': form,
        'coach': coach_user,
        'coach_profile': coach_profile,
        'LANGUAGE_CODE': translation.get_language(),
        'club': club
    })
    return render(request, 'club_dashboard/coachs/editCoach.html', context)


@club_permission_required('deleteCoach')
@login_required
def deleteCoach(request, id):
    """Deletes a coach from the club along with their products and services."""
    user = request.user

    # Get the club (either from director or administrator profile)
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    coach_profile = get_object_or_404(CoachProfile, id=id)
    coach = get_object_or_404(User, userprofile__Coach_profile=coach_profile)
    coach_name = coach.username

    # Delete all products created by this coach
    products_deleted = ProductsModel.objects.filter(creator=coach).delete()

    # Delete all services created by this coach
    services_deleted = ServicesModel.objects.filter(creator=coach).delete()

    # Delete coach profile and user account
    coach_profile.delete()
    coach.delete()

    # Send notification
    send_notification(
        user,
        club,
        f"تم حذف التاجر 🗑️ {coach_name} وجميع منتجاته وخدماته من المنصة"
    )

    messages.success(
        request,
        f"Employee and their {products_deleted[0]} products and {services_deleted[0]} services have been deleted successfully."
    )
    return redirect('viewCoachs')


from django.shortcuts import render, redirect
from django.utils import timezone, translation
from .forms import ArticleModelForm


@club_permission_required('addArticle')
def addArticle(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)
    form = ArticleModelForm()

    if request.method == 'POST':
        form = ArticleModelForm(data=request.POST, files=request.FILES)
        if form.is_valid():
            art = form.save(commit=False)
            art.club = club
            art.creator = user
            art.creation_date = timezone.now()
            art.save()
            return redirect('viewArticles')

    context['form'] = form
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/blog/addArticle.html', context)


# views.py - Updated editArticle view
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import translation
from django.contrib import messages


@club_permission_required('editArticle')
def editArticle(request, id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Use get_object_or_404 for better error handling
    art = get_object_or_404(Blog, id=id, club=club)

    if request.method == 'POST':
        form = ArticleModelForm(data=request.POST, files=request.FILES, instance=art)
        if form.is_valid():
            updated_article = form.save(commit=False)
            # Ensure the club and creator remain the same
            updated_article.club = club
            updated_article.creator = art.creator  # Keep original creator
            updated_article.save()

            # Add success message
            if translation.get_language() == 'ar':
                messages.success(request, 'تم تحديث المقال بنجاح!')
            else:
                messages.success(request, 'Article updated successfully!')

            return redirect('viewArticles')
        else:
            # Add error message if form is invalid
            if translation.get_language() == 'ar':
                messages.error(request, 'يرجى تصحيح الأخطاء أدناه.')
            else:
                messages.error(request, 'Please correct the errors below.')
    else:
        form = ArticleModelForm(instance=art)

    context = {
        'form': form,
        'article': art,  # Pass the article instance for additional template usage
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/blog/editArticle.html', context)


@club_permission_required('viewArticles')
def viewArticles(request):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    arts = Blog.objects.filter(club=club)  # Only show articles from this club

    # Get statistics for the dashboard
    total_articles = arts.count()
    current_month = timezone.now().month
    current_year = timezone.now().year
    new_articles_this_month = arts.filter(
        creation_date__month=current_month,
        creation_date__year=current_year
    ).count()

    # Get most popular articles (assuming you have a views or likes field)
    # If not, you can add this feature later
    popular_articles = arts.order_by('-id')[:3]

    context = {
        'arts': arts,
        'total_articles': total_articles,
        'new_articles_this_month': new_articles_this_month,
        'popular_articles': popular_articles.count(),
        'club': club,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/blog/viewArticless.html', context)


@club_permission_required('deleteArticle')
def DeleteArticle(request, id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    try:
        art = Blog.objects.get(id=id, club=club)
        art.delete()
    except Blog.DoesNotExist:
        pass

    return redirect('viewArticles')


@club_permission_required('viewDirectors')
def viewDirectors(request):
    context = {}
    user = request.user
    userprofile = getattr(user, 'userprofile', None)

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    if club:
        directors = UserProfile.objects.filter(
            account_type='2',
            director_profile__club=club
        ).select_related('user', 'director_profile')

        receptionists = UserProfile.objects.filter(
            account_type='5',
            receptionist_profile__club=club
        ).select_related('user', 'receptionist_profile')

        administrators = UserProfile.objects.filter(
            account_type='6',
            administrator_profile__club=club
        ).select_related('user', 'administrator_profile')

        accountants = UserProfile.objects.filter(
            account_type='7',
            accountant_profile__club=club
        ).select_related('user', 'accountant_profile')

        vendor_managers = UserProfile.objects.filter(
            account_type='8',
            vendor_manager_profile__club=club
        ).select_related('user', 'vendor_manager_profile')

        custom_role_staff = UserProfile.objects.filter(
            account_type='9',
            custom_role_profile__club=club
        ).select_related('user', 'custom_role_profile', 'custom_role')

        staff_list = []

        for director in directors:
            staff_list.append({
                'userprofile': director,
                'role': 'مدير عام',
                'role_en': 'General Manager',
                'profile': director.director_profile,
                'profile_type': 'director'
            })

        for receptionist in receptionists:
            staff_list.append({
                'userprofile': receptionist,
                'role': 'دعم فني',
                'role_en': 'technical support',
                'profile': receptionist.receptionist_profile,
                'profile_type': 'receptionist'
            })

        for administrator in administrators:
            staff_list.append({
                'userprofile': administrator,
                'role': 'إداري',
                'role_en': 'Administrator',
                'profile': administrator.administrator_profile,
                'profile_type': 'administrator'
            })

        for accountant in accountants:
            staff_list.append({
                'userprofile': accountant,
                'role': 'محاسب',
                'role_en': 'Accountant',
                'profile': accountant.accountant_profile,
                'profile_type': 'accountant'
            })

        for vendor_manager in vendor_managers:
            staff_list.append({
                'userprofile': vendor_manager,
                'role': 'مدير التجار',
                'role_en': 'Vendor Manager',
                'profile': vendor_manager.vendor_manager_profile,
                'profile_type': 'vendor_manager'
            })

        for staff in custom_role_staff:
            staff_list.append({
                'userprofile': staff,
                'role': staff.custom_role.name if staff.custom_role else 'Custom Role',
                'role_en': staff.custom_role.name if staff.custom_role else 'Custom Role',
                'profile': staff.custom_role_profile,
                'profile_type': 'custom_role'
            })

        staff_list.sort(key=lambda x: x['userprofile'].creation_date, reverse=True)

    else:
        staff_list = []

    context['LANGUAGE_CODE'] = translation.get_language()
    context['staff_list'] = staff_list
    return render(request, 'club_dashboard/directors/viewDirectors.html', context)


from club_dashboard.forms import VendorManagerProfileForm

import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.utils import translation
from accounts.models import UserProfile
from .forms import CustomRoleProfileForm


@club_permission_required('addDirector')
def addDirector(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or \
           getattr(user.userprofile.administrator_profile, 'club', None) or \
           getattr(user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        messages.error(request, "Unauthorized access.")
        return redirect('club_dashboard_index')

    ROLE_CHOICES = [
        ('2', 'مدير عام', 'General Manager'),
        ('5', 'دعم فني', 'Technical support'),
        ('6', 'إداري', 'Administrator'),
        ('7', 'محاسب', 'Accountant'),
        ('8', 'مدير التجار', 'Vendor Manager'),
        ('custom', 'دور مخصص', 'Custom Role'),
    ]

    # Get all custom roles for the club
    custom_roles = CustomRole.objects.filter(club=club, is_active=True)

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        full_name = request.POST.get('full_name')
        phone = request.POST.get('phone')
        about = request.POST.get('about')
        role = request.POST.get('role')
        custom_role_id = request.POST.get('custom_role')
        region = request.POST.get('region')

        # Basic validation
        if not username:
            messages.error(request, "Username is required.")
            return redirect('addDirector')

        if not full_name:
            messages.error(request, "Full name is required.")
            return redirect('addDirector')

        if not phone:
            messages.error(request, "Phone number is required.")
            return redirect('addDirector')

        if not password:
            messages.error(request, "Password is required.")
            return redirect('addDirector')

        # Check if username already exists
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('addDirector')

        # Email validation for specific roles
        roles_requiring_email = ['5', '6', '7', '8']
        if role in roles_requiring_email:
            if not email:
                messages.error(request, "Email is required for this role.")
                return redirect('addDirector')
            if User.objects.filter(email=email).exists():
                messages.error(request, "Email is already in use.")
                return redirect('addDirector')
        elif email and User.objects.filter(email=email).exists():
            messages.error(request, "Email is already in use.")
            return redirect('addDirector')

        # Custom role validation
        if role == 'custom' and not custom_role_id:
            messages.error(request, "Please select a custom role.")
            return redirect('addDirector')

        # Region validation for Vendor Manager
        if role == '8' and not region:
            messages.error(request, "Region is required for Vendor Manager role.")
            return redirect('addDirector')

        if role == 'custom' and not custom_role_id:
            messages.error(request, "Please select a custom role.")
            return redirect('addDirector')

        # Prepare form data
        form_data = {
            'full_name': full_name,
            'phone': phone,
            'about': about,
            'email': email if role in ['5', '6', '7', '8', 'custom'] else None
        }

        # Add about field for all roles except coaches (role 4)
        if role != '4' and about:
            form_data['about'] = about

        # Add region for Vendor Manager
        if role == '8' and region:
            form_data['region'] = region

        # Create appropriate form based on role
        form = None
        if role == '2':
            form = DirectorProfileForm(form_data)
        elif role == '5':
            form = ReceptionistProfileForm(form_data)
        elif role == '6':
            form = AdministratorProfileForm(form_data)
        elif role == '7':
            form = AccountantProfileForm(form_data)
        elif role == '8':
            form = VendorManagerProfileForm(form_data)
        elif role == 'custom':
            form = CustomRoleProfileForm(form_data)

        if form and form.is_valid():
            try:
                # Create user
                new_user = User.objects.create(
                    username=username,
                    email=email or '',
                    first_name=full_name.split()[0] if full_name else '',
                    last_name=' '.join(full_name.split()[1:]) if len(full_name.split()) > 1 else ''
                )
                new_user.set_password(password)
                new_user.save()

                # Create profile
                profile = form.save(commit=False)
                profile.club = club
                profile.save()

                # Determine account type based on role
                account_type_mapping = {
                    '2': '2',  # Director
                    '5': '5',  # Receptionist
                    '6': '6',  # Administrator
                    '7': '7',  # Accountant
                    '8': '8',  # Vendor Manager
                    'custom': '9',
                }

                user_profile_data = {
                    'user': new_user,
                    'account_type': account_type_mapping.get(role, '2'),
                }

                # Assign profile to appropriate field
                if role == '2':
                    user_profile_data['director_profile'] = profile
                elif role == '5':
                    user_profile_data['receptionist_profile'] = profile
                elif role == '6':
                    user_profile_data['administrator_profile'] = profile
                elif role == '7':
                    user_profile_data['accountant_profile'] = profile
                elif role == '8':
                    user_profile_data['vendor_manager_profile'] = profile
                elif role == 'custom':
                    user_profile_data['custom_role_profile'] = profile
                    custom_role = get_object_or_404(CustomRole, id=custom_role_id, club=club)
                    user_profile_data['custom_role'] = custom_role

                UserProfile.objects.create(**user_profile_data)

                # Success message
                role_names = {
                    '2': 'Director',
                    '5': 'Technical Support',
                    '6': 'Administrator',
                    '7': 'Accountant',
                    '8': 'Vendor Manager',
                    'custom': 'Custom Role'
                }
                messages.success(request, f"{role_names[role]} '{full_name}' added successfully.")
                return redirect('viewDirectors')

            except Exception as e:
                # If user was created but profile creation failed, clean up
                if 'new_user' in locals():
                    new_user.delete()
                messages.error(request, f"An error occurred while creating the staff member: {str(e)}")
                return redirect('addDirector')
        else:
            if form:
                # Display form errors
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
            else:
                messages.error(request, "Invalid role selected.")
            return redirect('addDirector')

    # Prepare context for GET request
    from accounts.fields import REGIONS_AND_CITIES
    regions = list(REGIONS_AND_CITIES.keys())
    context['regions'] = regions
    context['regions_json'] = json.dumps(REGIONS_AND_CITIES, ensure_ascii=False)
    context['LANGUAGE_CODE'] = translation.get_language()
    context['role_choices'] = ROLE_CHOICES
    context['custom_roles'] = custom_roles

    return render(request, 'club_dashboard/directors/addDirector.html', context)


from accounts.models import VendorManagerProfile, CustomRoleProfile


@club_permission_required('editDirector')
def editDirector(request, id, role):
    context = {}
    user = request.user

    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('club_dashboard_index')

    club = user.userprofile.director_profile.club

    ROLE_MAPPING = {
        '2': {
            'profile_model': DirectorProfile,
            'profile_field': 'director_profile',
            'form_class': DirectorProfileForm,
            'name': 'Director',
            'name_ar': 'مدير عام'
        },
        '5': {
            'profile_model': ReceptionistProfile,
            'profile_field': 'receptionist_profile',
            'form_class': ReceptionistProfileForm,
            'name': 'technical support',
            'name_ar': 'دعم فني'
        },
        '6': {
            'profile_model': AdministrativeProfile,
            'profile_field': 'administrator_profile',
            'form_class': AdministratorProfileForm,
            'name': 'Administrator',
            'name_ar': 'إداري'
        },
        '7': {
            'profile_model': AccountantProfile,
            'profile_field': 'accountant_profile',
            'form_class': AccountantProfileForm,
            'name': 'Accountant',
            'name_ar': 'محاسب'
        },
        '8': {
            'profile_model': VendorManagerProfile,
            'profile_field': 'vendor_manager_profile',
            'form_class': VendorManagerProfileForm,
            'name': 'Vendor Manager',
            'name_ar': 'مدير التجار'
        },
        '9': {
            'profile_model': CustomRoleProfile,
            'profile_field': 'custom_role_profile',
            'form_class': CustomRoleProfileForm,
            'name': 'Custom Role',
            'name_ar': 'دور مخصص'
        }
    }

    if role not in ROLE_MAPPING:
        messages.error(request, "Invalid role specified.")
        return redirect('viewDirectors')

    role_info = ROLE_MAPPING[role]

    try:
        profile = get_object_or_404(role_info['profile_model'], id=id, club=club)
        user_profile = get_object_or_404(UserProfile, **{role_info['profile_field']: profile})
        staff_user = user_profile.user
    except:
        messages.error(request, f"{role_info['name']} not found or unauthorized access.")
        return redirect('viewDirectors')

    form = role_info['form_class'](instance=profile)
    username = staff_user.username
    email = staff_user.email

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        form = role_info['form_class'](request.POST, instance=profile)

        if form.is_valid():
            if User.objects.exclude(id=staff_user.id).filter(username=username).exists():
                messages.error(request, "Username already exists.")
                return redirect('editDirector', id=id, role=role)

            if User.objects.exclude(id=staff_user.id).filter(email=email).exists():
                messages.error(request, "Email is already in use.")
                return redirect('editDirector', id=id, role=role)

            staff_user.username = username
            staff_user.email = email
            if password:
                staff_user.set_password(password)
            staff_user.save()

            updated_profile = form.save(commit=False)
            updated_profile.club = club
            updated_profile.save()

            messages.success(request, f"{role_info['name']} updated successfully.")
            return redirect('viewDirectors')
        else:
            messages.error(request, "Please correct the form errors.")

    context['LANGUAGE_CODE'] = translation.get_language()
    context['role'] = role
    context['role_name'] = role_info['name']
    context['role_name_ar'] = role_info['name_ar']

    return render(request, 'club_dashboard/directors/editDirector.html', {
        'form': form,
        'email': email,
        'username': username,
        'staff_user': staff_user,
        'profile': profile,
        'role': role,
        'role_name': role_info['name'],
        'role_name_ar': role_info['name_ar'],
        'LANGUAGE_CODE': context['LANGUAGE_CODE']
    })


@club_permission_required('deleteDirector')
def deleteDirector(request, id, role):
    user = request.user

    # ✅ Ensure the user is a director before proceeding
    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('viewDirectors')

    club = user.userprofile.director_profile.club

    # Role mapping for profile deletion
    ROLE_MAPPING = {
        '2': {
            'profile_model': DirectorProfile,
            'profile_field': 'director_profile',
            'name': 'Director',
            'name_ar': 'مدير عام'
        },
        '5': {
            'profile_model': ReceptionistProfile,
            'profile_field': 'receptionist_profile',
            'name': 'Receptionist',
            'name_ar': 'موظف استقبال'
        },
        '6': {
            'profile_model': AdministrativeProfile,
            'profile_field': 'administrator_profile',
            'name': 'Administrator',
            'name_ar': 'إداري'
        },
        '7': {
            'profile_model': AccountantProfile,
            'profile_field': 'accountant_profile',
            'name': 'Accountant',
            'name_ar': 'محاسب'
        },
        '8': {
            'profile_model': VendorManagerProfile,
            'profile_field': 'vendor_manager_profile',
            'name': 'Vendor Manager',
            'name_ar': 'مدير التجار'
        },
        '9': {
            'profile_model': CustomRoleProfile,
            'profile_field': 'custom_role_profile',
            'name': 'Custom Role',
            'name_ar': 'دور مخصص'
        }
    }

    # Validate role
    if role not in ROLE_MAPPING:
        messages.error(request, "Invalid role specified.")
        return redirect('viewDirectors')

    role_info = ROLE_MAPPING[role]

    # Get the profile and user based on role
    try:
        staff_profile = get_object_or_404(role_info['profile_model'], id=id, club=club)
        user_profile = get_object_or_404(UserProfile, **{role_info['profile_field']: staff_profile})
        staff_user = user_profile.user
    except:
        messages.error(request, f"{role_info['name']} not found or unauthorized access.")
        return redirect('viewDirectors')

    # ✅ Additional security check - ensure staff belongs to the same club
    if staff_profile.club != club:
        messages.error(request, f"You cannot delete a {role_info['name'].lower()} from another club.")
        return redirect('viewDirectors')

    # ✅ Prevent directors from deleting themselves
    if staff_user == user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('viewDirectors')

    # ✅ Delete the profile and user
    try:
        staff_profile.delete()
        staff_user.delete()
        messages.success(request, f"{role_info['name']} deleted successfully.")
    except Exception as e:
        messages.error(request, f"An error occurred while deleting the {role_info['name'].lower()}.")

    return redirect('viewDirectors')


@club_permission_required('viewClubNotifications')
def viewClubNotifications(request):
    context = {}
    """Displays all club notifications and marks them as read."""
    user = request.user

    # ✅ Ensure the user has a valid director profile
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # ✅ Fetch all notifications for the club
    notifications = Notification.objects.filter(club=club).order_by('-created_at')

    # ✅ Ensure only unread notifications are marked as read
    unread_count = notifications.filter(is_read=False).update(is_read=True)
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/notifications/viewClubNotifications.html', {
        'notifications': notifications,
        'unread_count': unread_count,  # ✅ Pass unread count for better UI
        'club': club,
    })


def delete_notification(request, notification_id):
    """Delete a specific notification"""
    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id)
            # Check if the notification belongs to the user's club
            club = getattr(request.user.userprofile.director_profile, 'club', None) or getattr(
                request.user.userprofile.administrator_profile, 'club', None) or getattr(
                user.userprofile.vendor_manager_profile, 'club', None)
            if notification.club == club:
                notification.delete()
                messages.success(request, "Notification deleted successfully.")
            else:
                messages.error(request, "You don't have permission to delete this notification.")
        except Notification.DoesNotExist:
            messages.error(request, "Notification not found.")

    return redirect('viewClubNotifications')


def delete_all_notifications(request):
    """Delete all notifications for the club"""
    if request.method == 'POST':
        club = getattr(request.user.userprofile.director_profile, 'club', None) or getattr(
            request.user.userprofile.administrator_profile, 'club', None) or getattr(
            user.userprofile.vendor_manager_profile, 'club', None)
        if club:
            deleted_count, _ = Notification.objects.filter(club=club).delete()
            messages.success(request, f"Deleted {deleted_count} notifications.")
        else:
            messages.error(request, "No club associated with your account.")

    return redirect('viewClubNotifications')


def mark_notifications_read(request):
    """Marks all unread notifications as read for the club."""
    user = request.user

    # ✅ Ensure the user has a valid director profile
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # ✅ Mark only unread notifications as read
    updated_count = Notification.objects.filter(club=club, is_read=False).update(is_read=True)

    return JsonResponse({'status': 'success', 'message': f'Marked {updated_count} notifications as read'})


def reviews_list(request):
    context = {}
    """
    Fetch all reviews for the club associated with the logged-in user.
    """
    # Assuming the logged-in user is a director or admin
    user = request.user

    try:
        # Get the club associated with the logged-in user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(
            user.userprofile.administrator_profile, 'club', None) or getattr(user.userprofile.vendor_manager_profile,
                                                                             'club', None)

        # Fetch all reviews for coaches in this club
        reviews = Review.objects.filter(coach__club=club).select_related(
            'student', 'coach'
        ).order_by('-created_at')
        # Calculate average rating for the club
        avg_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0
        context['LANGUAGE_CODE'] = translation.get_language()
        return render(request, 'club_dashboard/reviews_list.html', {
            'reviews': reviews,
            'avg_rating': avg_rating,
            'total_reviews': reviews.count(),
            'club': club
        })

    except AttributeError:
        messages.error(request, "لا يمكن العثور على المنصة الخاصة بك.")
        return redirect('club_dashboard')


@login_required
@club_permission_required('club_orders')
def club_orders(request):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Only show delivered, confirmed, completed, and cancelled orders to directors
    orders = Order.objects.filter(club=club).exclude(
        status__in=['pending', 'paid']
    ).order_by('-created_at')

    status_filter = request.GET.get('status')
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)

    # Rest of the view remains the same...
    from django.db.models import Count, Q, Case, When, IntegerField, Value, CharField, Sum, F, DecimalField

    orders = orders.annotate(
        has_products=Count('items', filter=Q(items__product__isnull=False)),
        has_services=Count('items', filter=Q(items__service__isnull=False))
    ).annotate(
        order_type=Case(
            When(has_products__gt=0, has_services=0, then=Value('products')),
            When(has_products=0, has_services__gt=0, then=Value('services')),
            When(has_products__gt=0, has_services__gt=0, then=Value('mixed')),
            default=Value('unknown'),
            output_field=CharField()
        ),
        order_type_display=Case(
            When(has_products__gt=0, has_services=0, then=Value('منتجات')),
            When(has_products=0, has_services__gt=0, then=Value('خدمات')),
            When(has_products__gt=0, has_services__gt=0, then=Value('منتجات وخدمات')),
            default=Value('غير معروف'),
            output_field=CharField()
        )
    )

    # Add commission summary statistics
    from django.db.models import Sum
    total_orders_value = orders.aggregate(total=Sum('total_price'))['total'] or 0
    total_commissions = orders.aggregate(total=Sum('total_vendor_commission'))['total'] or 0
    total_club_revenue = orders.aggregate(total=Sum('club_revenue'))['total'] or 0

    context = {
        'orders': orders,
        'status_filter': status_filter or 'all',
        'club': club,
        'commission_stats': {
            'total_orders_value': total_orders_value,
            'total_commissions': total_commissions,
            'total_club_revenue': total_club_revenue
        }
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/orders/club_orders.html', context)


@login_required
@club_permission_required('update_order_status')
def update_order_status(request, order_id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    if request.method == 'POST':
        try:
            print(f"Request to update order {order_id} - POST data: {request.POST}")

            order = get_object_or_404(Order, id=order_id, club=club)
            new_status = request.POST.get('status')

            # Handle cancellation with reasons
            if new_status == 'cancelled':
                return handle_order_cancellation(request, order, user)

            # Handle other status updates
            return handle_regular_status_update(request, order, new_status, user)

        except Order.DoesNotExist:
            print(f"Order {order_id} not found")
            return JsonResponse({'status': 'error', 'message': 'Order not found'}, status=404)
        except Exception as e:
            print(f"Unexpected error in update_order_status: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


def handle_order_cancellation(request, order, user):
    """Handle order cancellation with detailed reasons"""
    try:
        # Parse JSON data if it's sent as JSON
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST

        reason = data.get('cancellation_reason') or data.get('reason')
        custom_reason = data.get('custom_reason', '')
        additional_notes = data.get('additional_notes', '')

        if not reason:
            return JsonResponse({
                'status': 'error',
                'message': 'Cancellation reason is required'
            }, status=400)

        # Validate reason
        valid_reasons = [choice[0] for choice in OrderCancellation.CANCELLATION_REASONS]
        if reason not in valid_reasons:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid cancellation reason'
            }, status=400)

        # If reason is 'other', custom_reason is required
        if reason == 'other' and not custom_reason.strip():
            return JsonResponse({
                'status': 'error',
                'message': 'Custom reason is required when selecting "Other"'
            }, status=400)

        # Check if order can be cancelled
        if order.status in ['completed', 'cancelled']:
            return JsonResponse({
                'status': 'error',
                'message': f'Cannot cancel order with status: {order.get_status_display()}'
            }, status=400)

        old_status = order.status

        # Create cancellation record first
        cancellation = OrderCancellation.objects.create(
            order=order,
            reason=reason,
            custom_reason=custom_reason if reason == 'other' else '',
            additional_notes=additional_notes,
            cancelled_by=user
        )

        # Update order status
        order.status = 'cancelled'
        order.save()

        # Handle stock restoration for confirmed orders
        if old_status == 'confirmed':
            restore_product_stock(order)

        # Handle service cancellations
        handle_service_cancellation(order)

        # Create notification
        create_cancellation_notification(order, cancellation)

        print(f"Order {order.id} cancelled successfully by user {user.id}")

        return JsonResponse({
            'status': 'success',
            'message': f'تم إلغاء الطلب #{order.id} بنجاح',
            'cancellation_reason': cancellation.get_reason_display_text()
        })

    except Exception as e:
        print(f"Error in handle_order_cancellation: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'status': 'error',
            'message': f'Error cancelling order: {str(e)}'
        }, status=500)


def handle_regular_status_update(request, order, new_status, user):
    """Handle regular status updates (non-cancellation)"""
    try:
        print(f"Current order status: {order.status}, New status: {new_status}")

        if new_status not in dict(Order.STATUS_CHOICES):
            return JsonResponse({'status': 'error', 'message': 'Invalid status'}, status=400)

        old_status = order.status
        order.status = new_status

        try:
            order.full_clean()
            order.save()
            print(f"Order saved successfully with new status: {new_status}")
        except Exception as save_error:
            print(f"Error saving order: {save_error}")
            return JsonResponse({'status': 'error', 'message': f'Database error: {str(save_error)}'}, status=500)

        # Determine order type
        try:
            has_products = OrderItem.objects.filter(order=order, product__isnull=False).exists()
            has_services = OrderItem.objects.filter(order=order, service__isnull=False).exists()
            order_type = "mixed" if (has_products and has_services) else "products" if has_products else "services"
        except Exception:
            order_type = "unknown"

        # Handle confirmation logic
        if new_status == 'confirmed' and old_status == 'pending':
            process_order_confirmation(order)

        # Create notification
        create_status_update_notification(order, new_status, order_type)

        return JsonResponse({
            'status': 'success',
            'message': f'تم تحديث حالة الطلب إلى {dict(Order.STATUS_CHOICES)[new_status]}'
        })

    except Exception as e:
        print(f"Error in handle_regular_status_update: {str(e)}")
        return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'}, status=500)


def restore_product_stock(order):
    """Restore product stock when cancelling confirmed orders"""
    try:
        product_items = OrderItem.objects.filter(order=order, product__isnull=False)
        for item in product_items:
            product = item.product
            product.stock += item.quantity
            product.save()
            print(f"Restored product stock for {product.title}: {product.stock}")
    except Exception as e:
        print(f"Error restoring product stock: {str(e)}")


def handle_service_cancellation(order):
    """Handle service-related cancellations"""
    try:
        service_items = OrderItem.objects.filter(order=order, service__isnull=False)
        for item in service_items:
            service = item.service

            # Cancel active service orders
            try:
                service_orders = ServiceOrderModel.objects.filter(
                    service=service,
                    student=order.user,
                    is_complited=False
                )

                for service_order in service_orders:
                    service_order.is_complited = True  # Mark as completed (cancelled)
                    service_order.save()
                    print(f"Cancelled service order: {service_order.id}")

            except Exception as e:
                print(f"Error cancelling service orders: {str(e)}")

            # Handle appointment cancellations
            try:
                if hasattr(order.user, 'userprofile') and hasattr(order.user.userprofile, 'student_profile'):
                    student_profile = order.user.userprofile.student_profile

                    bookings = SalonBooking.objects.filter(
                        student=student_profile
                    ).select_related('appointment')

                    booking_services = BookingService.objects.filter(
                        service=service,
                        booking__in=bookings
                    )

                    for booking_service in booking_services:
                        if (booking_service.booking and
                                hasattr(booking_service.booking, 'appointment') and
                                booking_service.booking.appointment.is_paid):
                            appointment = booking_service.booking.appointment
                            appointment.is_paid = False  # Revert payment status
                            appointment.save()
                            print(f"Reverted payment for appointment ID {appointment.id}")

            except Exception as e:
                print(f"Error handling appointment cancellations: {str(e)}")

    except Exception as e:
        print(f"Error in handle_service_cancellation: {str(e)}")


import logging
from django.db import transaction
from django.utils import timezone
from django.contrib.auth.models import User

logger = logging.getLogger(__name__)


def process_order_confirmation(order):
    """
    Process order confirmation - update stock and subscriptions
    This should be called when an order status changes from 'pending' to 'confirmed'
    """
    try:
        with transaction.atomic():
            # Process product items - reduce stock
            product_items = order.items.filter(product__isnull=False)
            for item in product_items:
                product = item.product
                if product.stock >= item.quantity:
                    product.stock -= item.quantity
                    product.save()
                    logger.info(f"Updated product stock for {product.title}: {product.stock}")
                else:
                    raise ValueError(f"Insufficient stock for product {product.title}")

            # Process service items - update subscriptions
            service_items = order.items.filter(service__isnull=False)
            for item in service_items:
                service = item.service

                # Handle service subscriptions
                from students.models import ServiceOrderModel

                # Look for any active subscription for this user and service
                existing_service_order = ServiceOrderModel.objects.filter(
                    student=order.user,
                    service=service,
                    is_complited=False  # Only active subscriptions
                ).order_by('-end_datetime').first()

                if existing_service_order:
                    # Extend existing subscription
                    subscription_months = service.pricing_period_months * item.quantity

                    # If the current subscription is still active, extend from its end date
                    # Otherwise, extend from now
                    if existing_service_order.end_datetime > timezone.now():
                        new_end_datetime = existing_service_order.end_datetime + timezone.timedelta(
                            days=subscription_months * 30)
                    else:
                        new_end_datetime = timezone.now() + timezone.timedelta(days=subscription_months * 30)

                    # Update the existing subscription
                    existing_service_order.end_datetime = new_end_datetime
                    existing_service_order.price += service.price * item.quantity
                    existing_service_order.creation_date = timezone.now()
                    existing_service_order.is_complited = False  # Ensure it's still active
                    existing_service_order.save()

                    logger.info(
                        f"Extended existing subscription for service {service.title} (ID: {service.id}) until {new_end_datetime}")

                else:
                    # Create new service subscription
                    subscription_months = service.pricing_period_months * item.quantity
                    end_datetime = timezone.now() + timezone.timedelta(days=subscription_months * 30)

                    new_service_order = ServiceOrderModel.objects.create(
                        service=service,
                        student=order.user,
                        price=service.price * item.quantity,
                        is_complited=False,
                        end_datetime=end_datetime,
                        creation_date=timezone.now()
                    )

                    logger.info(
                        f"Created new subscription for service {service.title} (ID: {service.id}) until {end_datetime}")

                # Handle appointment payments
                try:
                    from receptionist_dashboard.models import SalonBooking, BookingService

                    # Find student profile
                    student_profile = None
                    if hasattr(order.user, 'userprofile') and hasattr(order.user.userprofile, 'student_profile'):
                        student_profile = order.user.userprofile.student_profile

                    if student_profile:
                        bookings = SalonBooking.objects.select_related('appointment')

                        booking_services = BookingService.objects.filter(
                            service=service,
                            booking__in=bookings
                        )

                        for booking_service in booking_services:
                            try:
                                if booking_service.booking and hasattr(booking_service.booking, 'appointment'):
                                    appointment = booking_service.booking.appointment
                                    if not appointment.is_paid:
                                        appointment.is_paid = True
                                        appointment.save()
                                        logger.info(f"Updated appointment ID {appointment.id} to paid")
                            except Exception as appt_error:
                                logger.error(f"Error updating appointment: {str(appt_error)}")

                except ImportError:
                    logger.warning("Salon booking models not available")
                except Exception as e:
                    logger.error(f"Error handling appointment payments: {str(e)}")

            # NEW: Process commission tracking
            process_order_commissions(order)

            # Update order commission fields
            order.update_commission_fields()

            logger.info(f"Successfully processed order confirmation for order {order.id}")
            return True

    except Exception as e:
        logger.error(f"Error processing order confirmation for order {order.id}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False


def process_order_commissions(order):
    """Process and save vendor commissions for confirmed order"""
    try:
        # Clear existing commission records
        OrderVendorCommission.objects.filter(order=order).delete()

        # Get commission breakdown
        breakdown = order.calculate_commission_breakdown()

        # Create commission records for each vendor
        for vendor_data in breakdown['vendor_breakdowns']:
            OrderVendorCommission.objects.create(
                order=order,
                vendor=vendor_data['vendor'],
                total_amount=vendor_data['total_amount'],
                commission_rate=vendor_data['commission_rate'],
                commission_amount=vendor_data['commission_amount']
            )

        logger.info(f"Created commission records for order {order.id}")

    except Exception as e:
        logger.error(f"Error processing commissions for order {order.id}: {str(e)}")


def create_cancellation_notification(order, cancellation):
    """Create notification for order cancellation"""
    try:
        # Determine order type for notification message
        has_products = OrderItem.objects.filter(order=order, product__isnull=False).exists()
        has_services = OrderItem.objects.filter(order=order, service__isnull=False).exists()

        if has_products and not has_services:
            order_type_ar = "المنتجات"
        elif has_services and not has_products:
            order_type_ar = "الخدمات"
        else:
            order_type_ar = "المنتجات والخدمات"

        reason_text = cancellation.get_reason_display_text()

        message = (f"تم إلغاء طلب {order_type_ar} رقم #{order.id}. "
                   f"سبب الإلغاء: {reason_text}. "
                   f"يرجى التواصل مع خدمة العملاء للمزيد من المعلومات.")

        Notification.objects.create(
            user=order.user,
            message=message,
            notification_type='order_cancelled'
        )
        print(f"Created cancellation notification for user {order.user.id}")

    except Exception as e:
        print(f"Error creating cancellation notification: {e}")


def create_status_update_notification(order, new_status, order_type):
    """Create notification for regular status updates"""
    try:
        if new_status == 'confirmed':
            if order_type == "products":
                message = f"تم تأكيد طلب المنتجات رقم #{order.id} وسيتم تجهيزه قريباً."
            elif order_type == "services":
                message = f"تم تأكيد طلب الخدمات رقم #{order.id} وسيتم تفعيلها قريباً."
            else:
                message = f"تم تأكيد طلب المنتجات والخدمات رقم #{order.id} وسيتم معالجته قريباً."
        elif new_status == 'completed':
            if order_type == "products":
                message = f"تم توصيل منتجاتك من الطلب رقم #{order.id} بنجاح. شكراً لك!"
            elif order_type == "services":
                message = f"تم تفعيل خدماتك من الطلب رقم #{order.id} بنجاح. شكراً لك!"
            else:
                message = f"تم اكتمال طلب المنتجات والخدمات رقم #{order.id} بنجاح. شكراً لك!"
        else:
            message = f"تم تحديث حالة الطلب رقم #{order.id} إلى {dict(Order.STATUS_CHOICES)[new_status]}."

        Notification.objects.create(
            user=order.user,
            message=message,
            notification_type='order_update'
        )
        print(f"Created notification for user {order.user.id}")

    except Exception as e:
        print(f"Error creating notification: {e}")


@login_required
def order_details_api(request, order_id):
    user = request.user
    print(f"طلب تفاصيل الطلب {order_id} بواسطة {user}")

    try:
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(
            user.userprofile.administrator_profile, 'club', None)

        order = Order.objects.get(id=order_id, club=club)
        order_items = OrderItem.objects.filter(order=order)

        # Get cancellation details if order is cancelled
        cancellation = None
        if order.status == 'cancelled':
            try:
                cancellation = OrderCancellation.objects.get(order=order)
            except OrderCancellation.DoesNotExist:
                print(f"No cancellation record found for cancelled order {order_id}")

        # Get commission breakdown
        commission_breakdown = None
        vendor_commissions = None
        if order.status == 'confirmed':
            commission_breakdown = order.calculate_commission_breakdown()
            # Convert Decimal to float for JSON serialization
            commission_breakdown['total_vendor_commission'] = float(commission_breakdown['total_vendor_commission'])
            commission_breakdown['club_revenue'] = float(commission_breakdown['club_revenue'])
            for vendor in commission_breakdown['vendor_breakdowns']:
                vendor['total_amount'] = float(vendor['total_amount'])
                vendor['commission_amount'] = float(vendor['commission_amount'])

            # Get vendor commission records
            vendor_commissions = OrderVendorCommission.objects.filter(order=order).select_related('vendor')

        context = {
            'order': order,
            'order_items': order_items,
            'cancellation': cancellation,
            'has_transfer_receipt': bool(order.transfer_receipt),
            'transfer_receipt_url': order.transfer_receipt.url if order.transfer_receipt else None,
            'commission_breakdown': commission_breakdown,
            'vendor_commissions': vendor_commissions,
        }
        context['LANGUAGE_CODE'] = translation.get_language()
        html = render_to_string('club_dashboard/orders/order_details_modal.html', context)

        return JsonResponse({
            'status': 'success',
            'html': html
        })
    except Order.DoesNotExist:
        print(f"الطلب {order_id} غير موجود")
        return JsonResponse({'status': 'error', 'message': 'Order not found'}, status=404)
    except Exception as e:
        print(f"خطأ غير متوقع: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def vendor_commission_report(request):
    """Generate vendor commission report"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Base queryset
    commissions = OrderVendorCommission.objects.filter(
        order__club=club,
        order__status='confirmed'
    ).select_related('vendor', 'order')

    # Apply date filters
    if start_date:
        commissions = commissions.filter(order__created_at__gte=start_date)
    if end_date:
        commissions = commissions.filter(order__created_at__lte=end_date)

    # Group by vendor
    from django.db.models import Sum, Count
    vendor_summary = commissions.values(
        'vendor__id',
        'vendor__business_name_en',
        'vendor__full_name'
    ).annotate(
        total_orders=Count('order', distinct=True),
        total_sales=Sum('total_amount'),
        total_commission=Sum('commission_amount'),
        avg_commission_rate=models.Avg('commission_rate')
    ).order_by('-total_commission')

    context = {
        'vendor_summary': vendor_summary,
        'commissions': commissions,
        'start_date': start_date,
        'end_date': end_date,
        'club': club
    }

    return render(request, 'club_dashboard/reports/vendor_commission_report.html', context)


from students.models import OrderVendorCommission
from accountant_dashboard.models import VATSettings


@login_required
def order_full_details(request, order_id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or \
           getattr(user.userprofile.administrator_profile, 'club', None) or \
           getattr(user.userprofile.vendor_manager_profile, 'club', None)

    try:
        # Get the order with optimized queries
        order = Order.objects.get(id=order_id, club=club)
        order_items = OrderItem.objects.filter(order=order).select_related(
            'product',
            'service',
            'product__creator__userprofile__Coach_profile',
            'service__creator__userprofile__Coach_profile'
        )

        # Get cancellation details if order is cancelled
        cancellation = None
        if order.status == 'cancelled':
            cancellation = OrderCancellation.objects.filter(order=order).first()

        # Get commission breakdown
        commission_breakdown = order.calculate_commission_breakdown()

        # Organize items by vendor
        vendor_items = {}
        for item in order_items:
            vendor = None
            if item.product and hasattr(item.product.creator.userprofile, 'Coach_profile'):
                vendor = item.product.creator.userprofile.Coach_profile
            elif item.service and hasattr(item.service.creator.userprofile, 'Coach_profile'):
                vendor = item.service.creator.userprofile.Coach_profile

            if vendor:
                if vendor.id not in vendor_items:
                    vendor_items[vendor.id] = {
                        'vendor': vendor,
                        'items': []
                    }
                vendor_items[vendor.id]['items'].append(item)

        # Add items to vendor breakdowns if they exist
        if 'vendor_breakdowns' in commission_breakdown:
            for vendor_data in commission_breakdown['vendor_breakdowns']:
                vendor_id = vendor_data['vendor'].id
                if vendor_id in vendor_items:
                    vendor_data['items'] = vendor_items[vendor_id]['items']

        context = {
            'order': order,
            'order_items': order_items,
            'cancellation': cancellation,
            'has_transfer_receipt': bool(order.transfer_receipt),
            'transfer_receipt_url': order.transfer_receipt.url if order.transfer_receipt else None,
            'commission_breakdown': commission_breakdown,
            'club': club,
            'LANGUAGE_CODE': translation.get_language(),
        }

        return render(request, 'club_dashboard/orders/order_full_details.html', context)

    except Order.DoesNotExist:
        messages.error(request, _('Order not found'))
        return redirect('club_orders')
    except Exception as e:
        messages.error(request, _('An error occurred while loading order details'))
        logger.error(f"Error in order_full_details: {str(e)}")
        return redirect('club_orders')


@login_required
def get_cancellation_details(request, order_id):
    """API endpoint to get cancellation details for an order"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    try:
        order = get_object_or_404(Order, id=order_id, club=club)

        if not hasattr(order, 'cancellation'):
            return JsonResponse({'status': 'error', 'message': 'Order is not cancelled'}, status=404)

        cancellation = order.cancellation

        return JsonResponse({
            'status': 'success',
            'cancellation': {
                'reason': cancellation.reason,
                'reason_display': cancellation.get_reason_display_text(),
                'custom_reason': cancellation.custom_reason,
                'additional_notes': cancellation.additional_notes,
                'cancelled_by': cancellation.cancelled_by.get_full_name() or cancellation.cancelled_by.username,
                'cancelled_at': cancellation.cancelled_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        })

    except Exception as e:
        print(f"Error getting cancellation details: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
@club_permission_required('club_financial_dashboard')
def club_financial_dashboard(request):
    user = request.user

    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('home')

    club = user.userprofile.director_profile.club

    # Date range handling
    start_date = request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

    time_period = request.GET.get('time_period', 'monthly')

    if not start_date:
        start_date = (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.now().strftime('%Y-%m-%d')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
    except ValueError:
        start_date_obj = timezone.now() - timezone.timedelta(days=30)
        end_date_obj = timezone.now()

    # Base queries
    orders = Order.objects.filter(
        club=club,
        created_at__gte=start_date_obj,
        created_at__lte=end_date_obj
    ).order_by('-created_at')

    # Sales Gross (all orders before deductions)
    sales_gross = orders.aggregate(Sum('total_price'))['total_price__sum'] or 0

    # Net Revenue (after commissions)
    net_revenue = orders.filter(status__in=['confirmed', 'completed']).aggregate(
        Sum('club_revenue'))['club_revenue__sum'] or 0

    # Payment method breakdown
    credit_card_orders = orders.filter(payment_method='credit_card')
    cash_orders = orders.filter(payment_method='cash_on_delivery')

    # Revenue calculations
    total_revenue = orders.filter(status__in=['confirmed', 'completed']).aggregate(
        Sum('total_price'))['total_price__sum'] or 0
    pending_revenue = orders.filter(status='pending').aggregate(
        Sum('total_price'))['total_price__sum'] or 0

    credit_card_revenue = credit_card_orders.filter(status__in=['confirmed', 'completed']).aggregate(
        Sum('total_price'))['total_price__sum'] or 0
    cash_revenue = cash_orders.filter(status__in=['confirmed', 'completed']).aggregate(
        Sum('total_price'))['total_price__sum'] or 0

    # Cancellation and refund metrics
    cancelled_orders = orders.filter(status='cancelled')
    cancellation_rate = 0
    if orders.count() > 0:
        cancellation_rate = round((cancelled_orders.count() / orders.count()) * 100, 2)

    refunds = RefundDispute.objects.filter(
        deal__club=club,
        status__in=['approved', 'resolved'],
        created_at__gte=start_date_obj,
        created_at__lte=end_date_obj
    )
    refund_amount = refunds.aggregate(Sum('approved_refund_amount'))['approved_refund_amount__sum'] or 0
    refund_rate = 0
    if total_revenue > 0:
        refund_rate = round((refund_amount / total_revenue) * 100, 2)

    # Time period data (weekly/monthly/daily)
    time_period_data = []
    if time_period == 'weekly':
        # Weekly breakdown
        current_date = start_date_obj
        while current_date <= end_date_obj:
            week_start = current_date
            week_end = current_date + timezone.timedelta(days=6)

            week_orders = Order.objects.filter(
                club=club,
                created_at__gte=week_start,
                created_at__lte=week_end
            )

            week_sales_gross = week_orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
            week_net_revenue = week_orders.filter(status__in=['confirmed', 'completed']).aggregate(
                Sum('club_revenue'))['club_revenue__sum'] or 0

            time_period_data.append({
                'period': f"Week {week_start.isocalendar()[1]}",
                'start_date': week_start.strftime('%Y-%m-%d'),
                'end_date': week_end.strftime('%Y-%m-%d'),
                'sales_gross': float(week_sales_gross),
                'net_revenue': float(week_net_revenue),
                'credit_card': float(week_orders.filter(payment_method='credit_card').aggregate(
                    Sum('total_price'))['total_price__sum'] or 0),
                'cash': float(week_orders.filter(payment_method='cash_on_delivery').aggregate(
                    Sum('total_price'))['total_price__sum'] or 0)
            })

            current_date = week_end + timezone.timedelta(days=1)

    elif time_period == 'daily':
        # Daily breakdown
        current_date = start_date_obj
        while current_date <= end_date_obj:
            day_orders = Order.objects.filter(
                club=club,
                created_at__date=current_date.date()
            )

            day_sales_gross = day_orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
            day_net_revenue = day_orders.filter(status__in=['confirmed', 'completed']).aggregate(
                Sum('club_revenue'))['club_revenue__sum'] or 0

            time_period_data.append({
                'period': current_date.strftime('%Y-%m-%d'),
                'sales_gross': float(day_sales_gross),
                'net_revenue': float(day_net_revenue),
                'credit_card': float(day_orders.filter(payment_method='credit_card').aggregate(
                    Sum('total_price'))['total_price__sum'] or 0),
                'cash': float(day_orders.filter(payment_method='cash_on_delivery').aggregate(
                    Sum('total_price'))['total_price__sum'] or 0)
            })

            current_date += timezone.timedelta(days=1)

    else:  # monthly
        # Monthly breakdown (default)
        for i in range(6):
            month_date = timezone.now() - timezone.timedelta(days=30 * i)
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if month_date.month == 12:
                month_end = month_date.replace(year=month_date.year + 1, month=1, day=1, hour=0, minute=0, second=0,
                                               microsecond=0) - timezone.timedelta(seconds=1)
            else:
                month_end = month_date.replace(month=month_date.month + 1, day=1, hour=0, minute=0, second=0,
                                               microsecond=0) - timezone.timedelta(seconds=1)

            month_orders = Order.objects.filter(
                club=club,
                created_at__gte=month_start,
                created_at__lte=month_end
            )

            month_sales_gross = month_orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
            month_net_revenue = month_orders.filter(status__in=['confirmed', 'completed']).aggregate(
                Sum('club_revenue'))['club_revenue__sum'] or 0

            time_period_data.append({
                'period': month_start.strftime('%B %Y'),
                'sales_gross': float(month_sales_gross),
                'net_revenue': float(month_net_revenue),
                'credit_card': float(month_orders.filter(payment_method='credit_card').aggregate(
                    Sum('total_price'))['total_price__sum'] or 0),
                'cash': float(month_orders.filter(payment_method='cash_on_delivery').aggregate(
                    Sum('total_price'))['total_price__sum'] or 0)
            })

        time_period_data.reverse()
        time_period_data = [item for item in time_period_data if item['sales_gross'] > 0 or item['net_revenue'] > 0]

    # Category performance
    # Replace the top_activity_types query with this corrected version:

    from collections import defaultdict

    # Get all completed order items
    order_items = OrderItem.objects.filter(
        order__club=club,
        order__status__in=['confirmed', 'completed'],
        order__created_at__gte=start_date_obj,
        order__created_at__lte=end_date_obj
    ).select_related(
        'service__creator__userprofile__Coach_profile__activity_type',
        'product__creator__userprofile__Coach_profile__activity_type'
    )

    # Aggregate by activity type
    activity_revenue = defaultdict(float)
    activity_counts = defaultdict(int)

    for item in order_items:
        if item.service and item.service.creator:
            if hasattr(item.service.creator, 'userprofile'):
                coach_profile = item.service.creator.userprofile.Coach_profile
                if coach_profile and coach_profile.activity_type:
                    activity_revenue[coach_profile.activity_type.name] += float(item.price)
                    activity_counts[coach_profile.activity_type.name] += 1
        elif item.product and item.product.creator:
            # For products, get activity type from the creator's coach profile
            if hasattr(item.product.creator, 'userprofile'):
                coach_profile = item.product.creator.userprofile.Coach_profile
                if coach_profile and coach_profile.activity_type:
                    activity_revenue[coach_profile.activity_type.name] += float(item.price)
                    activity_counts[coach_profile.activity_type.name] += 1

    # Convert to the format needed for the template
    top_categories = [
        {
            'name': activity_type,
            'revenue': revenue,
            'sales_count': activity_counts[activity_type]
        }
        for activity_type, revenue in sorted(
            activity_revenue.items(),
            key=lambda x: x[1],
            reverse=True
        )
    ][:5]

    order_items = OrderItem.objects.filter(
        order__club=club,
        order__status__in=['confirmed', 'completed']
    ).select_related('product', 'service')

    category_performance = {}
    product_sales = {}
    service_sales = {}

    for item in order_items:
        if item.product:
            # Product category tracking
            classifications = item.product.classification.all()
            for classification in classifications:
                if classification.title in category_performance:
                    category_performance[classification.title]['product_count'] += item.quantity
                    category_performance[classification.title]['revenue'] += float(item.price * item.quantity)
                else:
                    category_performance[classification.title] = {
                        'type': 'product',
                        'product_count': item.quantity,
                        'revenue': float(item.price * item.quantity)
                    }

            # Individual product tracking
            product_id = item.product.id
            product_name = item.product.title
            if product_id in product_sales:
                product_sales[product_id]['quantity'] += item.quantity
                product_sales[product_id]['revenue'] += float(item.price * item.quantity)
            else:
                product_sales[product_id] = {
                    'name': product_name,
                    'quantity': item.quantity,
                    'revenue': float(item.price * item.quantity)
                }

        if item.service:
            # Service category tracking
            classifications = item.service.classification.all()
            for classification in classifications:
                if classification.title in category_performance:
                    category_performance[classification.title]['service_count'] += item.quantity
                    category_performance[classification.title]['revenue'] += float(item.price * item.quantity)
                else:
                    category_performance[classification.title] = {
                        'type': 'service',
                        'service_count': item.quantity,
                        'revenue': float(item.price * item.quantity)
                    }

            # Individual service tracking
            service_id = item.service.id
            service_name = item.service.title
            if service_id in service_sales:
                service_sales[service_id]['quantity'] += item.quantity
                service_sales[service_id]['revenue'] += float(item.price * item.quantity)
            else:
                service_sales[service_id] = {
                    'name': service_name,
                    'quantity': item.quantity,
                    'revenue': float(item.price * item.quantity)
                }

    # Sort and limit results

    top_products = sorted(
        [v for k, v in product_sales.items()],
        key=lambda x: x['revenue'], reverse=True
    )[:5]

    top_services = sorted(
        [v for k, v in service_sales.items()],
        key=lambda x: x['revenue'], reverse=True
    )[:5]

    # Vendor commission breakdown
    vendor_commissions = OrderVendorCommission.objects.filter(
        order__club=club,
        order__created_at__gte=start_date_obj,
        order__created_at__lte=end_date_obj
    ).values('vendor__full_name', 'vendor__business_name_ar').annotate(
        total_sales=Sum('total_amount'),
        total_commission=Sum('commission_amount'),
        order_count=Count('order')
    ).order_by('-total_sales')[:5]

    print("Time period data:", time_period_data)

    # Prepare months data for the monthly chart
    months_data = []
    for i in range(6):
        month_date = timezone.now() - timezone.timedelta(days=30 * i)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timezone.timedelta(seconds=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timezone.timedelta(seconds=1)

        month_orders = Order.objects.filter(
            club=club,
            created_at__gte=month_start,
            created_at__lte=month_end
        )

        months_data.append({
            'month': month_start.strftime('%B %Y'),
            'total': float(
                month_orders.filter(status__in=['confirmed', 'completed'])
                .aggregate(Sum('total_price'))['total_price__sum'] or 0
            ),
            'credit_card': float(
                month_orders.filter(payment_method='credit_card', status__in=['confirmed', 'completed'])
                .aggregate(Sum('total_price'))['total_price__sum'] or 0
            ),
            'cash': float(
                month_orders.filter(payment_method='cash_on_delivery', status__in=['confirmed', 'completed'])
                .aggregate(Sum('total_price'))['total_price__sum'] or 0
            )
        })

    months_data.reverse()

    print(f"Querying from {start_date} to {end_date}")
    print(f"Club ID: {club.id}")
    print(f"Initial order count: {orders.count()}")

    context = {
        'sales_gross': sales_gross,
        'net_revenue': net_revenue,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'credit_card_revenue': credit_card_revenue,
        'cash_revenue': cash_revenue,
        'credit_card_percentage': round((credit_card_revenue / total_revenue * 100) if total_revenue > 0 else 0),
        'cash_percentage': round((cash_revenue / total_revenue * 100) if total_revenue > 0 else 0),
        'cancellation_rate': cancellation_rate,
        'refund_rate': refund_rate,
        'refund_amount': refund_amount,
        'start_date': start_date,
        'end_date': end_date,
        'time_period': time_period,
        'time_period_data': json.dumps(time_period_data),
        'months_data': json.dumps(months_data),
        'top_categories': json.dumps(top_categories),
        'all_orders': orders,
        'cancelled_orders': cancelled_orders,
        'vendor_commissions': vendor_commissions,
        'refunds': refunds,
        'LANGUAGE_CODE': translation.get_language()
    }

    return render(request, 'club_dashboard/financial/dashboard.html', context)


from django.http import HttpResponse
import csv
from datetime import datetime


@login_required
def export_financial_data(request):
    user = request.user
    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('home')

    club = user.userprofile.director_profile.club

    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
    except (ValueError, TypeError):
        start_date_obj = timezone.now() - timezone.timedelta(days=30)
        end_date_obj = timezone.now()

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{start_date}_to_{end_date}.csv"'

    writer = csv.writer(response)

    # Write header row
    if translation.get_language() == 'ar':
        writer.writerow([
            'رقم الطلب', 'التاريخ', 'حالة الطلب', 'طريقة الدفع',
            'إجمالي المبيعات', 'عمولة المنصة', 'صافي الإيرادات',
            'نوع الطلب', 'عدد العناصر', 'ملاحظات'
        ])
    else:
        writer.writerow([
            'Order ID', 'Date', 'Status', 'Payment Method',
            'Sales Gross', 'Platform Commission', 'Net Revenue',
            'Order Type', 'Item Count', 'Notes'
        ])

    # Get orders data
    orders = Order.objects.filter(
        club=club,
        created_at__gte=start_date_obj,
        created_at__lte=end_date_obj
    ).order_by('-created_at')

    # Helper function to safely get display values
    def safe_get_display(obj, field_name):
        """Safely get display value for a field"""
        try:
            # Try the get_*_display method first
            display_method = getattr(obj, f'get_{field_name}_display', None)
            if display_method and callable(display_method):
                return display_method()
            # Fall back to the raw field value
            return str(getattr(obj, field_name, ''))
        except:
            return ''

    # Write data rows
    for order in orders:
        if translation.get_language() == 'ar':
            writer.writerow([
                order.id,
                order.created_at.strftime('%Y-%m-%d'),
                safe_get_display(order, 'status'),
                safe_get_display(order, 'payment_method'),
                order.total_price,
                order.total_vendor_commission,
                order.club_revenue,
                safe_get_display(order, 'order_type'),
                order.items.count(),
                order.notes or ''
            ])
        else:
            writer.writerow([
                order.id,
                order.created_at.strftime('%Y-%m-%d'),
                safe_get_display(order, 'status'),
                safe_get_display(order, 'payment_method'),
                order.total_price,
                order.total_vendor_commission,
                order.club_revenue,
                safe_get_display(order, 'order_type'),
                order.items.count(),
                order.notes or ''
            ])

    return response


@login_required
def view_director_profile(request):
    """View the director's profile"""
    try:
        # Get the current user's UserProfile
        userprofile = request.user.userprofile

        # Check if user is a director
        if not userprofile.director_profile:
            return HttpResponseForbidden("You don't have permission to view this page")

        director = userprofile.director_profile

        context = {
            'director': director,
            'userprofile': userprofile,
        }
        context['LANGUAGE_CODE'] = translation.get_language()
        return render(request, 'accounts/profiles/Director/ViewDirectorProfile.html', context)
    except UserProfile.DoesNotExist:
        return HttpResponseForbidden("User profile not found")


@login_required
def edit_director_profile(request):
    """Edit the director's profile"""
    try:
        # Get the current user's UserProfile
        userprofile = request.user.userprofile

        # Check if user is a director
        if not userprofile.director_profile:
            return HttpResponseForbidden("You don't have permission to edit this page")

        director = userprofile.director_profile

        if request.method == 'POST':
            form = DirectorProfileForm(request.POST, request.FILES, instance=director)
            if form.is_valid():
                # Save director profile form
                director = form.save()

                # Handle profile image upload
                if 'profile_image_base64' in request.FILES:
                    image_file = request.FILES['profile_image_base64']
                    # Convert the image to base64
                    encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                    image_data = f"data:image/{image_file.content_type.split('/')[-1]};base64,{encoded_image}"

                    # Update the user profile
                    userprofile.profile_image_base64 = image_data
                    userprofile.save()

                return redirect(reverse('view_director_profile'))
        else:
            form = DirectorProfileForm(instance=director)

        context = {
            'form': form,
            'director': director,
            'userprofile': userprofile,
        }
        context['LANGUAGE_CODE'] = translation.get_language()
        return render(request, 'accounts/settings/Director/EditDirectorProfile.html', context)
    except UserProfile.DoesNotExist:
        return HttpResponseForbidden("User profile not found")


def handle_uploaded_image(image_file):
    """Convert uploaded image to base64 string"""
    if not image_file:
        return None

    # Check file size (limit to 2MB)
    if image_file.size > 2 * 1024 * 1024:
        raise ValueError("Image file too large (max 2MB)")

    # Get file extension
    file_extension = image_file.name.split('.')[-1].lower()
    if file_extension not in ['jpg', 'jpeg', 'png', 'gif']:
        raise ValueError("Unsupported file format")

    # Convert to base64
    encoded_image = base64.b64encode(image_file.read()).decode('utf-8')

    # Create data URL based on file type
    if file_extension in ['jpg', 'jpeg']:
        return f"data:image/jpeg;base64,{encoded_image}"
    elif file_extension == 'png':
        return f"data:image/png;base64,{encoded_image}"
    elif file_extension == 'gif':
        return f"data:image/gif;base64,{encoded_image}"


def toggle_dashboard_counts(request):
    """Toggle dashboard counts visibility - accessible only to staff"""
    if request.method == 'POST':
        settings = DashboardSettings.get_settings()
        settings.show_employee_client_counts = not settings.show_employee_client_counts
        settings.save()

        status = "shown" if settings.show_employee_client_counts else "hidden"

        if request.headers.get('Accept') == 'application/json':
            return JsonResponse({
                'success': True,
                'status': settings.show_employee_client_counts,
                'message': f"Dashboard count cards are now {status}!"
            })
        else:
            messages.success(request, f"Dashboard count cards are now {status}!")
            return redirect('/admin/')

    return redirect('/admin/')


def update_club_descriptions(request, club_id):
    club = get_object_or_404(ClubsModel, pk=club_id)
    if request.method == 'POST':
        club.productsDescription = request.POST.get('productsDescription', '')
        club.articlesDescription = request.POST.get('articlesDescription', '')
        club.save()
        messages.success(request, "Descriptions updated successfully.")
    return redirect('club_dashboard_index')


# def update_club_pricing(request, club_id):
#     club = get_object_or_404(ClubsModel, pk=club_id)
#     if request.method == 'POST':
#         new_pricing = []
#         i = 1
#         while f'name_{i}' in request.POST:
#             name = request.POST.get(f'name_{i}')
#             price = request.POST.get(f'price_{i}')
#             features = request.POST.get(f'features_{i}', '')
#             features_list = [f.strip() for f in features.split(',') if f.strip()]
#             new_pricing.append({
#                 'name': name,
#                 'price': price,
#                 'features': features_list
#             })
#             i += 1
#         club.pricing = new_pricing
#         club.save()
#         messages.success(request, "تم تحديث التسعير بنجاح." if request.LANGUAGE_CODE == 'ar' else "Pricing updated successfully.")
#     return redirect('club_dashboard_index')

def vendor_status(request, vendor_id):
    """Show vendor application status"""
    vendor = get_object_or_404(CoachProfile, id=vendor_id)
    return render(request, 'accounts/vendor_status.html', {'vendor': vendor})


from django.contrib.auth.decorators import login_required


@login_required
def vendor_approval_list(request):
    """List of vendors pending approval - only for directors"""
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.account_type != '2':  # Not a director
            messages.error(request, "غير مسموح لك بالوصول إلى هذه الصفحة.")
            return redirect('dashboard')

        # Get director's club
        director_profile = user_profile.director_profile
        if not director_profile or not director_profile.club:
            messages.error(request, "لم يتم العثور على منصة مرتبطة بحسابك.")
            return redirect('dashboard')

        # Get pending vendors for this club
        pending_vendors = CoachProfile.objects.filter(
            club=director_profile.club,
            approval_status='pending'
        ).order_by('-created_at')

        return render(request, 'accounts/vendor_approval_list.html', {
            'pending_vendors': pending_vendors,
            'club': director_profile.club
        })

    except UserProfile.DoesNotExist:
        messages.error(request, "ملف المستخدم غير موجود.")
        return redirect('signin')


from accounts.forms import VendorApprovalForm
from django.core.mail import send_mail
from django.conf import settings


def send_vendor_approval_email(vendor, approved=True):
    """Send email to vendor about approval/rejection"""
    try:
        if approved:
            subject = f"تم قبول طلبك - {vendor.business_name_en}"
            message = f"""
            مرحباً {vendor.full_name},

            تم قبول طلب تسجيلك كبائع في منصة {vendor.club.name}.

            سيتم إنشاء حسابك قريباً وسيتم إرسال بيانات الدخول إليك.

            شكراً لانضمامك إلينا!
            """
        else:
            subject = f"طلب التسجيل - {vendor.business_name_en}"
            message = f"""
            مرحباً {vendor.full_name},

            نأسف لإبلاغك أنه لم يتم قبول طلب تسجيلك كبائع في منصة {vendor.club.name}.

            {vendor.approval_notes if vendor.approval_notes else ''}

            شكراً لك على اهتمامك.
            """

        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [vendor.email],
            fail_silently=True,
        )
    except Exception as e:
        print(f"Error sending approval email: {e}")


@login_required
def vendor_approval_action(request, vendor_id):
    """Handle vendor approval/rejection actions"""
    try:
        vendor = CoachProfile.objects.get(id=vendor_id)
        user_profile = UserProfile.objects.get(user=request.user)

        if user_profile.account_type != '2':  # Only directors can approve
            messages.error(request, "You don't have permission to perform this action.")
            return redirect('vendor_approval_list')

        if request.method == 'POST':
            action = request.POST.get('action')
            notes = request.POST.get('notes', '')

            if action == 'approve':
                vendor.approve(request.user, notes)
                messages.success(request, f"Vendor {vendor.full_name} approved successfully.")
            elif action == 'reject':
                vendor.reject(request.user, notes)
                messages.success(request, f"Vendor {vendor.full_name} rejected.")
            else:
                messages.error(request, "Invalid action.")

            return redirect('vendor_approval_list')

    except CoachProfile.DoesNotExist:
        messages.error(request, "Vendor not found.")
        return redirect('vendor_approval_list')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('vendor_approval_list')


from django.views.generic import DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from accounts.models import CoachProfile


class VendorApprovalDetailView(LoginRequiredMixin, DetailView):
    model = CoachProfile
    template_name = 'club_dashboard/vendor_approval_detail.html'
    context_object_name = 'vendor'

    def get_queryset(self):
        # Only show pending vendors for the director's club
        user = self.request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or \
               getattr(user.userprofile.administrator_profile, 'club', None) or \
               getattr(user.userprofile.vendor_manager_profile, 'club', None)

        if club:
            return CoachProfile.objects.filter(approval_status='pending', club=club)
        return CoachProfile.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vendor = self.object

        # Add comprehensive vendor information
        context.update({
            'vendor_files': self.get_vendor_files(vendor),
            'vendor_branches': vendor.get_branches_display(),
            'business_phone_numbers': vendor.business_phone_numbers or [],
            'working_hours': vendor.working_hours or {},
            'special_hours': vendor.special_hours or {},
            'subcategories': vendor.subcategories.all(),
            'activity_type': vendor.activity_type,
            'commission_info': self.get_commission_info(vendor),
            'LANGUAGE_CODE': translation.get_language(),
        })
        return context

    def get_vendor_files(self, vendor):
        """Get all uploaded files from the vendor"""
        files = {}

        # Commercial registration certificate
        if vendor.commercial_registration_certificate:
            files['commercial_registration'] = {
                'name': 'شهادة السجل التجاري' if translation.get_language() == 'ar' else 'Commercial Registration Certificate',
                'data': vendor.commercial_registration_certificate,
                'type': 'document'
            }

        # Tax certificate
        if vendor.tax_certificate:
            files['tax_certificate'] = {
                'name': 'شهادة الرقم الضريبي' if translation.get_language() == 'ar' else 'Tax Certificate',
                'data': vendor.tax_certificate,
                'type': 'document'
            }

        # Business document
        if vendor.business_document_file:
            files['business_document'] = {
                'name': f"وثيقة النشاط التجاري ({vendor.get_business_document_type_display()})" if translation.get_language() == 'ar' else f'Business Document ({vendor.get_business_document_type_display()})',
                'data': vendor.business_document_file,
                'type': 'document'
            }

        # Store logo
        if vendor.store_logo_base64:
            files['store_logo'] = {
                'name': 'شعار المتجر' if translation.get_language() == 'ar' else 'Store Logo',
                'data': vendor.store_logo_base64,
                'type': 'image'
            }

        # Business photo
        if vendor.business_photo_base64:
            files['business_photo'] = {
                'name': 'صورة النشاط التجاري' if translation.get_language() == 'ar' else 'Business Photo',
                'data': vendor.business_photo_base64,
                'type': 'image'
            }

        # Profile image
        if vendor.profile_image_base64:
            files['profile_image'] = {
                'name': 'صورة الملف الشخصي' if translation.get_language() == 'ar' else 'Profile Image',
                'data': vendor.profile_image_base64,
                'type': 'image'
            }

        return files

    def get_commission_info(self, vendor):
        """Get commission information for the vendor"""
        from club_dashboard.models import Commission

        commission = Commission.objects.filter(
            club=vendor.club,
            commission_type='vendor',
            vendor_classification=vendor.vendor_classification,
            is_active=True
        ).first()

        return commission

    def post(self, request, *args, **kwargs):
        """Handle vendor approval actions"""
        vendor = self.get_object()
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')

        if action == 'approve':
            return self.approve_vendor(vendor, notes)
        elif action == 'reject_with_feedback':
            return self.reject_with_feedback(vendor, notes)
        elif action == 'delete':
            return self.delete_vendor(vendor)
        else:
            messages.error(request, 'Invalid action selected.')
            return redirect('vendor_approval_detail', pk=vendor.pk)

    def approve_vendor(self, vendor, notes):
        """Approve the vendor and create user account"""
        print("🔹 approve_vendor called")
        print("🔹 Vendor ID:", vendor.id)
        print("🔹 Vendor email:", vendor.email)
        print("🔹 Notes:", notes)

        try:
            print("🔹 Entering transaction.atomic()")
            with transaction.atomic():

                # Create user account with fixed password
                print("🔹 Creating User account...")
                user = User.objects.create_user(
                    username=vendor.email,
                    email=vendor.email,
                    password='12345678'
                )
                print("✅ User created:", user.id, user.username)

                # Create user profile and link to existing CoachProfile
                print("🔹 Creating UserProfile...")
                user_profile = UserProfile.objects.create(
                    user=user,
                    account_type='4',  # Coach/Vendor
                    Coach_profile=vendor  # Link to the existing CoachProfile
                )
                print("✅ UserProfile created:", user_profile.id)
                print("✅ UserProfile linked to CoachProfile:", vendor.id)

                # Update vendor profile
                print("🔹 Updating vendor approval fields...")
                vendor.approval_status = 'approved'
                vendor.approved_at = timezone.now()
                vendor.approved_by = self.request.user
                vendor.approval_notes = notes
                vendor.save()
                print("✅ Vendor updated: approval_status =", vendor.approval_status)

                # Assign commission
                print("🔹 Assigning commission...")
                vendor.assign_commission()
                print("✅ Commission assigned")

                # Send approval email
                print("🔹 Sending approval email...")
                self.send_approval_email(vendor, user)
                print("✅ Approval email sent")

                messages.success(
                    self.request,
                    f'Vendor {vendor.full_name} has been approved successfully!'
                )

        except Exception as e:
            print("🔥 Error in approve_vendor:", str(e))
            messages.error(self.request, f'Error approving vendor: {str(e)}')

        print("🔹 Redirecting to vendor_approval_list")
        return redirect('vendor_approval_list')

    def reject_with_feedback(self, vendor, notes):
        """Reject vendor but keep pending for resubmission"""
        try:
            # Update vendor with rejection notes but keep status as pending
            vendor.approval_notes = notes
            vendor.save()

            # Send rejection email with edit link
            self.send_rejection_email(vendor, notes)

            messages.success(self.request,
                             f'Rejection feedback sent to {vendor.full_name}. They can resubmit after making changes.')

        except Exception as e:
            messages.error(self.request, f'Error sending rejection feedback: {str(e)}')

        return redirect('vendor_approval_list')

    def delete_vendor(self, vendor):
        """Permanently delete the vendor application"""
        try:
            vendor_name = vendor.full_name
            vendor.delete()
            messages.success(self.request, f'Vendor application for {vendor_name} has been permanently deleted.')

        except Exception as e:
            messages.error(self.request, f'Error deleting vendor: {str(e)}')

        return redirect('vendor_approval_list')

    def send_approval_email(self, vendor, user):
        """Send approval email with login credentials"""
        from django.core.mail import send_mail
        from django.conf import settings

        # The password is fixed as '12345678' as set in the approve_vendor method
        password = '12345678'

        subject = f'تم قبول طلبك كبائع في {vendor.club.name}' if translation.get_language() == 'ar' else f'Your vendor application has been approved for {vendor.club.name}'

        if translation.get_language() == 'ar':
            message = f"""
            مرحباً {vendor.full_name},

            نحن سعداء لإبلاغك أنه تم قبول طلب تسجيلك كبائع في منصة {vendor.club.name}.

            بيانات تسجيل الدخول:
            اسم المستخدم: {user.username}
            كلمة المرور: {password}

            يمكنك الآن تسجيل الدخول وبدء إضافة منتجاتك وخدماتك.

            {vendor.approval_notes if vendor.approval_notes else ''}

            مرحباً بك في فريقنا!
            """
        else:
            message = f"""
            Hello {vendor.full_name},

            We are pleased to inform you that your vendor application for {vendor.club.name} has been approved.

            Login credentials:
            Username: {user.username}
            Password: {password}

            You can now log in and start adding your products and services.

            {vendor.approval_notes if vendor.approval_notes else ''}

            Welcome to our team!
            """

        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [vendor.email],
                fail_silently=False,
            )
        except Exception as e:
            print(f"Error sending approval email: {e}")

    def send_rejection_email(self, vendor, notes):
        """Send rejection email with feedback and edit link"""
        from django.core.mail import send_mail
        from django.conf import settings
        from django.urls import reverse

        # Generate edit link (you'll need to create this URL)
        edit_url = self.request.build_absolute_uri(
            reverse('vendor_edit_application', kwargs={'vendor_id': vendor.id})
        )

        subject = f'ملاحظات على طلب التسجيل في {vendor.club.name}' if translation.get_language() == 'ar' else f'Feedback on your application for {vendor.club.name}'

        if translation.get_language() == 'ar':
            message = f"""
            مرحباً {vendor.full_name},

            شكراً لك على اهتمامك بالانضمام إلى منصة {vendor.club.name} كبائع.

            بعد مراجعة طلبك، لدينا بعض الملاحظات التي نود منك تعديلها:

            {notes}

            يمكنك تعديل بياناتك وإعادة تقديم الطلب من خلال الرابط التالي:
            {edit_url}

            نتطلع لرؤية طلبك المحدث قريباً.

            تحياتنا
            فريق {vendor.club.name}
            """
        else:
            message = f"""
            Hello {vendor.full_name},

            Thank you for your interest in joining {vendor.club.name} as a vendor.

            After reviewing your application, we have some feedback that we'd like you to address:

            {notes}

            You can edit your information and resubmit your application using the following link:
            {edit_url}

            We look forward to seeing your updated application soon.

            Best regards,
            {vendor.club.name} Team
            """

        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [vendor.email],
                fail_silently=False,
            )
        except Exception as e:
            print(f"Error sending rejection email: {e}")


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Count, Q
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta
from .models import Category, SubCategory
from .forms import CategoryForm, SubCategoryForm


@club_permission_required('category_list')
@login_required
def category_list(request):
    """View to display all categories and subcategories"""
    # Get search query
    search_query = request.GET.get('search', '')

    # Filter categories based on search
    categories = Category.objects.all().order_by('name_en')
    if search_query:
        categories = categories.filter(
            Q(name_en__icontains=search_query) |
            Q(name_ar__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Annotate with subcategory count
    categories = categories.annotate(subcategory_count=Count('subcategories'))

    # Get subcategories
    subcategories = SubCategory.objects.select_related('category').order_by('name_en')
    if search_query:
        subcategories = subcategories.filter(
            Q(name_en__icontains=search_query) |
            Q(name_ar__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__name_en__icontains=search_query) |
            Q(category__name_ar__icontains=search_query)
        )

    # Pagination for categories
    category_paginator = Paginator(categories, 10)
    category_page = request.GET.get('category_page')
    categories = category_paginator.get_page(category_page)

    # Pagination for subcategories
    subcategory_paginator = Paginator(subcategories, 10)
    subcategory_page = request.GET.get('subcategory_page')
    subcategories = subcategory_paginator.get_page(subcategory_page)

    # Statistics
    stats = {
        'total_categories': Category.objects.count(),
        'active_categories': Category.objects.filter(is_active=True).count(),
        'total_subcategories': SubCategory.objects.count(),
        'active_subcategories': SubCategory.objects.filter(is_active=True).count(),
    }

    context = {
        'categories': categories,
        'subcategories': subcategories,
        'stats': stats,
        'search_query': search_query,
    }

    return render(request, 'club_dashboard/categories/category_list.html', context)


@login_required
def add_category(request):
    """View to add a new category"""
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            category = form.save()
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CategoryForm()

    context = {
        'form': form,
        'title': 'Add New Category',
        'submit_text': 'Create Category',
    }
    return render(request, 'club_dashboard/categories/category_form.html', context)


@login_required
def edit_category(request, category_id):
    """View to edit an existing category"""
    category = get_object_or_404(Category, id=category_id)

    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" has been updated successfully!')
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CategoryForm(instance=category)

    context = {
        'form': form,
        'category': category,
        'title': f'Edit Category - {category.name}',
        'submit_text': 'Update Category',
    }
    return render(request, 'club_dashboard/categories/category_form.html', context)


@login_required
def delete_category(request, category_id):
    """View to delete a category"""
    category = get_object_or_404(Category, id=category_id)

    # Check if category has subcategories
    if category.subcategories.exists():
        messages.error(request,
                       f'Cannot delete category "{category.name}" because it has subcategories. Please delete or move the subcategories first.')
        return redirect('category_list')

    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Category "{category_name}" has been deleted successfully!')
        return redirect('category_list')

    context = {
        'category': category,
        'subcategory_count': category.subcategories.count(),
    }
    return render(request, 'club_dashboard/categories/category_delete.html', context)


@login_required
def add_subcategory(request):
    """View to add a new subcategory"""
    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            subcategory = form.save()
            messages.success(request, f'Subcategory "{subcategory.name}" has been created successfully!')
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SubCategoryForm()

    context = {
        'form': form,
        'title': 'Add New Subcategory',
        'submit_text': 'Create Subcategory',
    }
    return render(request, 'club_dashboard/categories/subcategory_form.html', context)


@login_required
def edit_subcategory(request, subcategory_id):
    """View to edit an existing subcategory"""
    subcategory = get_object_or_404(SubCategory, id=subcategory_id)

    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES, instance=subcategory)
        if form.is_valid():
            subcategory = form.save()
            messages.success(request, f'Subcategory "{subcategory.name}" has been updated successfully!')
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SubCategoryForm(instance=subcategory)

    context = {
        'form': form,
        'subcategory': subcategory,
        'title': f'Edit Subcategory - {subcategory.name}',
        'submit_text': 'Update Subcategory',
    }
    return render(request, 'club_dashboard/categories/subcategory_form.html', context)


@login_required
def delete_subcategory(request, subcategory_id):
    """View to delete a subcategory"""
    subcategory = get_object_or_404(SubCategory, id=subcategory_id)

    if request.method == 'POST':
        subcategory_name = subcategory.name
        category_name = subcategory.category.name
        subcategory.delete()
        messages.success(request,
                         f'Subcategory "{subcategory_name}" from "{category_name}" has been deleted successfully!')
        return redirect('category_list')

    context = {
        'subcategory': subcategory,
    }
    return render(request, 'club_dashboard/categories/subcategory_delete.html', context)


@login_required
def category_detail(request, category_id):
    """View to show category details with its subcategories"""
    category = get_object_or_404(Category, id=category_id)
    subcategories = category.subcategories.all().order_by('name')

    # Pagination for subcategories
    paginator = Paginator(subcategories, 12)
    page = request.GET.get('page')
    subcategories = paginator.get_page(page)

    context = {
        'category': category,
        'subcategories': subcategories,
    }
    return render(request, 'club_dashboard/categories/category_detail.html', context)


# AJAX Views for dynamic functionality
@login_required
def get_subcategories(request, category_id):
    """AJAX view to get subcategories for a specific category"""
    category = get_object_or_404(Category, id=category_id)
    subcategories = category.subcategories.filter(is_active=True).values('id', 'name')
    return JsonResponse({'subcategories': list(subcategories)})


@login_required
def toggle_category_status(request, category_id):
    """AJAX view to toggle category active status"""
    if request.method == 'POST':
        category = get_object_or_404(Category, id=category_id)
        category.is_active = not category.is_active
        category.save()

        status_text = 'activated' if category.is_active else 'deactivated'
        messages.success(request, f'Category "{category.name}" has been {status_text}!')

        return redirect('category_list')

    return redirect('category_list')


@login_required
def toggle_subcategory_status(request, subcategory_id):
    """AJAX view to toggle subcategory active status"""
    if request.method == 'POST':
        subcategory = get_object_or_404(SubCategory, id=subcategory_id)
        subcategory.is_active = not subcategory.is_active
        subcategory.save()

        status_text = 'activated' if subcategory.is_active else 'deactivated'
        messages.success(request, f'Subcategory "{subcategory.name}" has been {status_text}!')

        return redirect('category_list')

    return redirect('category_list')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone, translation
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from .models import ProductsModel, CoachProfile, ClubsModel, ProductImg
from .forms import ProductApprovalForm


@login_required
@club_permission_required('manage_products')
def manage_products(request):
    """View to manage all products with approval status"""
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None)

    # Get filter parameters
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '')

    # Base queryset for products in this club
    products = ProductsModel.objects.filter(club=club).select_related(
        'creator', 'creator__userprofile', 'creator__userprofile__Coach_profile'
    ).prefetch_related('product_images', 'classification')

    # Apply status filter
    if status_filter == 'pending':
        products = products.filter(approval_status='pending')
    elif status_filter == 'approved':
        products = products.filter(approval_status='approved')
    elif status_filter == 'rejected':
        products = products.filter(approval_status='rejected')

    # Apply search filter
    if search_query:
        products = products.filter(
            Q(title__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__business_name_en__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__full_name__icontains=search_query) |
            Q(creator__email__icontains=search_query)
        )

    # Order by creation date (newest first)
    products = products.order_by('-creation_date')

    # Get statistics
    stats = {
        'total': ProductsModel.objects.filter(club=club).count(),
        'pending': ProductsModel.objects.filter(club=club, approval_status='pending').count(),
        'approved': ProductsModel.objects.filter(club=club, approval_status='approved').count(),
        'rejected': ProductsModel.objects.filter(club=club, approval_status='rejected').count(),
    }

    # Get latest pending products for quick review
    latest_pending = ProductsModel.objects.filter(
        club=club,
        approval_status='pending'
    ).select_related(
        'creator', 'creator__userprofile', 'creator__userprofile__Coach_profile'
    ).order_by('-creation_date')[:3]

    # Pagination
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context.update({
        'products': page_obj,
        'stats': stats,
        'latest_pending': latest_pending,
        'status_filter': status_filter,
        'search_query': search_query,
        'club': club,
        'LANGUAGE_CODE': translation.get_language()
    })

    return render(request, 'club_dashboard/products/manage_products.html', context)


@login_required
def pending_products(request):
    """View to show all pending products"""
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None)

    # Get search query if any
    search_query = request.GET.get('search', '')

    # Get all pending products for this club
    products = ProductsModel.objects.filter(
        club=club,
        approval_status='pending'
    ).select_related(
        'creator', 'creator__userprofile', 'creator__userprofile__Coach_profile'
    ).prefetch_related('product_images', 'classification')

    # Apply search filter if exists
    if search_query:
        products = products.filter(
            Q(title__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__business_name_en__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__full_name__icontains=search_query) |
            Q(creator__email__icontains=search_query)
        )

    # Order by creation date (newest first)
    products = products.order_by('-creation_date')

    # Pagination
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context.update({
        'products': page_obj,
        'search_query': search_query,
        'club': club,
        'LANGUAGE_CODE': translation.get_language()
    })

    return render(request, 'club_dashboard/products/pending_products.html', context)


@login_required
def approve_product(request, product_id):
    """Approve a product"""
    product = get_object_or_404(ProductsModel, id=product_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Check permissions
    if product.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذا المنتج')
        return redirect('manage_products')

    if request.method == 'POST':
        notes = request.POST.get('approval_notes', '')

        # Update product status
        product.approval_status = 'approved'
        product.approved_at = timezone.now()
        product.approved_by = user
        product.approval_notes = notes
        product.is_enabled = True
        product.save()

        # Send email notification to vendor
        try:
            vendor_profile = product.creator.userprofile.Coach_profile
            subject = f"تم قبول منتجك: {product.title}"
            context = {
                'product': product,
                'vendor': vendor_profile,
                'notes': notes,
                'club': club
            }
            html_message = render_to_string('emails/product_approved.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[product.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم قبول المنتج "{product.title}" بنجاح')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم قبول المنتج بنجاح'})

    return redirect('manage_products')


@login_required
def reject_product(request, product_id):
    """Reject a product"""
    product = get_object_or_404(ProductsModel, id=product_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Check permissions
    if product.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذا المنتج')
        return redirect('manage_products')

    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')

        # Update product status
        product.approval_status = 'rejected'
        product.approved_at = timezone.now()
        product.approved_by = user
        product.approval_notes = rejection_reason
        product.is_enabled = False
        product.save()

        # Send email notification to vendor
        try:
            vendor_profile = product.creator.userprofile.Coach_profile
            subject = f"تم رفض منتجك: {product.title}"
            context = {
                'product': product,
                'vendor': vendor_profile,
                'rejection_reason': rejection_reason,
                'club': club
            }
            html_message = render_to_string('products/emails/product_rejected.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[product.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم رفض المنتج "{product.title}"')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم رفض المنتج'})

    return redirect('manage_products')


@login_required
def product_detail(request, product_id):
    """View detailed information about a product"""
    product = get_object_or_404(ProductsModel, id=product_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Check permissions
    if product.club != club:
        messages.error(request, 'غير مسموح لك بعرض هذا المنتج')
        return redirect('manage_products')

    # Get product images
    product_images = product.product_images.all()

    # Get number of orders for this product
    order_count = OrderItem.objects.filter(product=product).count()

    context = {
        'product': product,
        'product_images': product_images,
        'vendor_profile': product.creator.userprofile.Coach_profile if hasattr(product.creator,
                                                                               'userprofile') else None,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
        'order_count': order_count
    }

    return render(request, 'club_dashboard/products/product_detail.html', context)


@login_required
def bulk_approve_products(request):
    """Bulk approve multiple products"""
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(
            user.userprofile.administrator_profile, 'club', None) or getattr(user.userprofile.vendor_manager_profile,
                                                                             'club', None)

        if not product_ids:
            messages.error(request, 'لم يتم تحديد أي منتجات')
            return redirect('manage_products')

        # Update products
        updated_count = ProductsModel.objects.filter(
            id__in=product_ids,
            club=club,
            approval_status='pending'
        ).update(
            approval_status='approved',
            approved_at=timezone.now(),
            approved_by=user,
            is_enabled=True
        )

        messages.success(request, f'تم قبول {updated_count} منتج بنجاح')

    return redirect('manage_products')


@login_required
def bulk_reject_products(request):
    """Bulk reject multiple products"""
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        rejection_reason = request.POST.get('bulk_rejection_reason', 'تم الرفض بواسطة الإدارة')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(
            user.userprofile.administrator_profile, 'club', None) or getattr(user.userprofile.vendor_manager_profile,
                                                                             'club', None)

        if not product_ids:
            messages.error(request, 'لم يتم تحديد أي منتجات')
            return redirect('manage_products')

        # Update products
        updated_count = ProductsModel.objects.filter(
            id__in=product_ids,
            club=club,
            approval_status='pending'
        ).update(
            approval_status='rejected',
            approved_at=timezone.now(),
            approved_by=user,
            approval_notes=rejection_reason,
            is_enabled=False
        )

        messages.success(request, f'تم رفض {updated_count} منتج')

    return redirect('manage_products')


def delete_product(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(ProductsModel, id=product_id)
        # Add any permission checks here
        product.delete()
        messages.success(request, 'Product deleted successfully')
        return redirect('manage_products')


# club_dashboard/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from .models import Commission, VendorCommissionAssignment
from .forms import CommissionForm
from accounts.models import CoachProfile

from accounts.models import CoachProfile


@login_required
@club_permission_required('commission_list')
def commission_list(request):
    """List all commissions with filtering and pagination"""
    # Get filter parameters
    commission_type = request.GET.get('type', '')
    classification = request.GET.get('classification', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    user = request.user

    # Get the club
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Base queryset
    commissions = Commission.objects.filter(club=club)

    # Apply filters
    if commission_type:
        commissions = commissions.filter(commission_type=commission_type)

    if classification:
        commissions = commissions.filter(vendor_classification=classification)

    if status == 'active':
        commissions = commissions.filter(is_active=True)
    elif status == 'inactive':
        commissions = commissions.filter(is_active=False)

    if search:
        commissions = commissions.filter(
            Q(name__icontains=search) |
            Q(commission_rate__icontains=search)
        )

    latest_vendor_commissions = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).order_by('-created_at')[:4]

    # Pagination
    paginator = Paginator(commissions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get unique vendor classifications for the club
    vendor_classifications = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).values_list('vendor_classification', flat=True).distinct()

    # Create choices list for the filter form
    classification_choices = [(c, c.capitalize()) for c in vendor_classifications]

    # Statistics
    stats = {
        'total_classifications': Commission.objects.filter(club=club).values(
            'vendor_classification').distinct().count(),
        'total_offers': Commission.objects.filter(club=club, commission_type='time_period').distinct().count(),
        # Assuming you have an Offer model
        'total_vendors': CoachProfile.objects.filter(club=club).count(),  # Assuming you have a Vendor model
        'active_commissions': Commission.objects.filter(club=club, is_active=True).count(),
    }

    context = {
        'page_obj': page_obj,
        'stats': stats,
        'current_filters': {
            'type': commission_type,
            'classification': classification,
            'status': status,
            'search': search,
        },
        'commission_types': Commission.get_commission_choices_with_all(),
        'vendor_classifications': classification_choices,  # Updated to use dynamic classifications
        'latest_vendor_commissions': latest_vendor_commissions,
    }

    return render(request, 'club_dashboard/commissions/list.html', context)


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages


@login_required
def commission_create(request):
    """Create a new commission"""

    print("DEBUG: commission_create view called")

    user = request.user
    print(f"DEBUG: User: {user}, Authenticated: {user.is_authenticated}")

    # Try getting the club
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)
    print(f"DEBUG: Club determined from user profile: {club}")

    if request.method == 'POST':
        print("DEBUG: Request method is POST")
        form = CommissionForm(request.POST, club=club)
        print(f"DEBUG: Form initialized with POST data and club: {form}")

        if form.is_valid():
            print("DEBUG: Form is valid")
            commission = form.save(commit=False)
            commission.club = club
            commission.created_by = user
            commission.save()
            print(f"DEBUG: Commission created and saved: {commission}")

            messages.success(request, 'تم إنشاء العمولة بنجاح')
            return redirect('commission_list')
        else:
            print("DEBUG: Form is not valid")
            print(f"DEBUG: Form errors: {form.errors}")

    else:
        print("DEBUG: Request method is GET")
        form = CommissionForm(club=club)
        print("DEBUG: Form initialized with empty data and club")

    return render(request, 'club_dashboard/commissions/create.html', {'form': form})


@login_required
def commission_edit(request, commission_id):
    """Edit an existing commission"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id,
                                   club=getattr(user.userprofile.director_profile, 'club', None) or getattr(
                                       user.userprofile.administrator_profile, 'club', None) or getattr(
                                       user.userprofile.vendor_manager_profile, 'club', None))

    if request.method == 'POST':
        form = CommissionForm(request.POST, instance=commission,
                              club=getattr(user.userprofile.director_profile, 'club', None) or getattr(
                                  user.userprofile.administrator_profile, 'club', None) or getattr(
                                  user.userprofile.vendor_manager_profile, 'club', None))
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث العمولة بنجاح')
            return redirect('commission_list')
    else:
        form = CommissionForm(instance=commission,
                              club=getattr(user.userprofile.director_profile, 'club', None) or getattr(
                                  user.userprofile.administrator_profile, 'club', None) or getattr(
                                  user.userprofile.vendor_manager_profile, 'club', None))

    return render(request, 'club_dashboard/commissions/edit.html', {
        'form': form,
        'commission': commission
    })


@login_required
@require_http_methods(["POST"])
def commission_delete(request, commission_id):
    """Delete a commission"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id,
                                   club=getattr(user.userprofile.director_profile, 'club', None) or getattr(
                                       user.userprofile.administrator_profile, 'club', None) or getattr(
                                       user.userprofile.vendor_manager_profile, 'club', None))

    # Check if commission is assigned to any vendors
    assigned_vendors_count = VendorCommissionAssignment.objects.filter(commission=commission).count()

    if assigned_vendors_count > 0:
        messages.error(request, f'لا يمكن حذف هذه العمولة لأنها مخصصة لـ {assigned_vendors_count} بائع')
    else:
        commission_name = commission.get_display_name()
        commission.delete()
        messages.success(request, f'تم حذف العمولة "{commission_name}" بنجاح')

    return redirect('commission_list')


@login_required
@require_http_methods(["POST"])
def commission_toggle_status(request, commission_id):
    """Toggle commission active/inactive status"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id,
                                   club=getattr(user.userprofile.director_profile, 'club', None) or getattr(
                                       user.userprofile.administrator_profile, 'club', None) or getattr(
                                       user.userprofile.vendor_manager_profile, 'club', None))

    commission.is_active = not commission.is_active
    commission.save()

    status_text = 'مفعلة' if commission.is_active else 'معطلة'
    messages.success(request, f'تم تغيير حالة العمولة إلى {status_text}')

    return redirect('commission_list')


@login_required
def commission_detail(request, commission_id):
    """View commission details and assigned vendors"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id,
                                   club=getattr(user.userprofile.director_profile, 'club', None) or getattr(
                                       user.userprofile.administrator_profile, 'club', None) or getattr(
                                       user.userprofile.vendor_manager_profile, 'club', None))

    # Get assigned vendors for vendor type commissions
    assigned_vendors = []
    if commission.commission_type == 'vendor':
        assigned_vendors = VendorCommissionAssignment.objects.filter(
            commission=commission
        ).select_related('vendor')

    # Get active time period commissions (for reference)
    active_time_commissions = Commission.objects.filter(
        club=getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                                 'club', None),
        commission_type='time_period',
        is_active=True,
        start_date__lte=timezone.now().date(),
        end_date__gte=timezone.now().date()
    )

    context = {
        'commission': commission,
        'assigned_vendors': assigned_vendors,
        'active_time_commissions': active_time_commissions,
    }

    return render(request, 'club_dashboard/commissions/detail.html', context)


@login_required
def vendor_commission_management(request):
    """Manage vendor commission assignments"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Get all approved vendors for the club
    vendors = CoachProfile.objects.filter(
        club=club,
        approval_status='approved'
    ).select_related('commission_assignment__commission')

    # Get available vendor commissions
    available_commissions = Commission.objects.filter(
        club=club,
        commission_type='vendor',
        is_active=True
    )

    # Get unique vendor classifications for the club
    vendor_classifications = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).values_list('vendor_classification', flat=True).distinct()

    # Create stats for each classification
    classification_stats = []
    for classification in vendor_classifications:
        count = CoachProfile.objects.filter(
            club=club,
            approval_status='approved',
            vendor_classification=classification
        ).count()
        classification_stats.append((classification, classification.capitalize(), count))

    # Handle bulk assignment
    if request.method == 'POST':
        vendor_id = request.POST.get('vendor_id')
        commission_id = request.POST.get('commission_id')

        if vendor_id and commission_id:
            vendor = get_object_or_404(CoachProfile, id=vendor_id, club=club)
            commission = get_object_or_404(Commission, id=commission_id, club=club)

            # Update or create assignment
            assignment, created = VendorCommissionAssignment.objects.get_or_create(
                vendor=vendor,
                defaults={'commission': commission}
            )

            if not created:
                assignment.commission = commission
                assignment.save()

            # Update vendor classification
            vendor.vendor_classification = commission.vendor_classification
            vendor.save()

            action = 'تم تخصيص' if created else 'تم تحديث'
            messages.success(request, f'{action} العمولة للبائع {vendor.full_name} بنجاح')

    # Pagination
    paginator = Paginator(vendors, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'available_commissions': available_commissions,
        'vendor_classifications': classification_stats,  # Updated to use dynamic stats
    }

    return render(request, 'club_dashboard/commissions/vendor_management.html', context)


@login_required
def commission_analytics(request):
    """Commission analytics and reports"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Get unique vendor classifications for the club
    vendor_classifications = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).values_list('vendor_classification', flat=True).distinct()

    # Vendor commission analytics
    vendor_commission_data = []
    for classification in vendor_classifications:
        commission = Commission.objects.filter(
            club=club,
            commission_type='vendor',
            vendor_classification=classification,
            is_active=True
        ).first()

        vendor_count = CoachProfile.objects.filter(
            club=club,
            vendor_classification=classification,
            approval_status='approved'
        ).count()

        vendor_commission_data.append({
            'classification': classification.capitalize(),
            'commission_rate': commission.commission_rate if commission else 0,
            'vendor_count': vendor_count,
            'total_potential_commission': (commission.commission_rate * vendor_count) if commission else 0
        })

    # Time period commissions
    time_period_commissions = Commission.objects.filter(
        club=club,
        commission_type='time_period',
        is_active=True
    ).order_by('start_date')

    # Current active time commission
    current_time_commission = Commission.get_time_period_commission(club)

    context = {
        'vendor_commission_data': vendor_commission_data,
        'time_period_commissions': time_period_commissions,
        'current_time_commission': current_time_commission,
        'total_vendors': CoachProfile.objects.filter(club=club, approval_status='approved').count(),
    }

    return render(request, 'club_dashboard/commissions/analytics.html', context)


def delete_commission(request, commission_id):
    if request.method == 'POST':
        commission = get_object_or_404(Commission, id=commission_id)
        # Add any permission checks here
        commission.delete()
        messages.success(request, 'Commission deleted successfully')
        return redirect('commission_list')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone, translation
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from accounts.models import CoachProfile, ClubsModel
from .forms import ServiceApprovalForm
from students.models import ServicesModel


@login_required
@club_permission_required('manage_services')
def manage_services(request):
    """View to manage all services with approval status"""
    context = {}
    user = request.user

    # More robust club detection
    club = None
    user_profile = getattr(user, 'userprofile', None)
    if user_profile:
        club = (getattr(user_profile, 'director_profile', None) and user_profile.director_profile.club or
                getattr(user_profile, 'administrator_profile', None) and user_profile.administrator_profile.club or
                getattr(user_profile, 'vendor_manager_profile', None) and user_profile.vendor_manager_profile.club)

    if not club:
        messages.error(request, 'غير مسموح لك بالوصول لهذه الصفحة')
        return redirect('club_dashboard')

    # Get filter parameters with validation
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '').strip()

    # Base queryset with optimized database queries
    services = ServicesModel.objects.filter(club=club).select_related(
        'creator__userprofile__Coach_profile'
    ).prefetch_related('classification')

    # Apply filters more efficiently
    status_filters = {
        'pending': services.filter(approval_status='pending'),
        'approved': services.filter(approval_status='approved'),
        'rejected': services.filter(approval_status='rejected'),
        'all': services
    }

    services = status_filters.get(status_filter, services)

    # Optimized search
    if search_query:
        services = services.filter(
            Q(title__icontains=search_query) |
            Q(desc__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__business_name_en__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__full_name__icontains=search_query) |
            Q(creator__email__icontains=search_query)
        ).distinct()

    # Get statistics in single query for better performance
    from django.db.models import Count
    stats = services.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(approval_status='pending')),
        approved=Count('id', filter=Q(approval_status='approved')),
        rejected=Count('id', filter=Q(approval_status='rejected'))
    )

    # Latest pending services
    latest_pending = services.filter(approval_status='pending')[:3]

    # Pagination
    paginator = Paginator(services.order_by('-creation_date'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context.update({
        'services': page_obj,
        'stats': stats,
        'latest_pending': latest_pending,
        'status_filter': status_filter,
        'search_query': search_query,
        'club': club,
        'LANGUAGE_CODE': translation.get_language()
    })

    return render(request, 'club_dashboard/services/manage_services.html', context)


@login_required
@club_permission_required('manage_services')  # Add this decorator for consistency
def pending_services(request):
    """View to show all pending services in a dedicated page"""
    user = request.user
    club = get_user_club(user)  # Use the helper function

    if not club:
        messages.error(request, 'غير مسموح لك بالوصول لهذه الصفحة')
        return redirect('club_dashboard')

    # Get and clean search query
    search_query = request.GET.get('search', '').strip()

    # Base queryset for pending services with optimized database queries
    services = ServicesModel.objects.filter(
        club=club,
        approval_status='pending'
    ).select_related(
        'creator__userprofile__Coach_profile'
    ).prefetch_related('classification')

    # Apply search filter if provided
    if search_query:
        services = services.filter(
            Q(title__icontains=search_query) |
            Q(desc__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__business_name_en__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__full_name__icontains=search_query) |
            Q(creator__email__icontains=search_query)
        ).distinct()

    # Get statistics in a single query for better performance
    from django.db.models import Count, Q
    stats = ServicesModel.objects.filter(club=club).aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(approval_status='pending')),
        approved=Count('id', filter=Q(approval_status='approved')),
        rejected=Count('id', filter=Q(approval_status='rejected'))
    )

    # Order by creation date (newest first) and paginate
    services = services.order_by('-creation_date')

    # Pagination
    paginator = Paginator(services, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'services': page_obj,
        'stats': stats,
        'search_query': search_query,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
        'is_pending_page': True
    }

    return render(request, 'club_dashboard/services/pending_services.html', context)


@login_required
def approve_service(request, service_id):
    """Approve a service"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذه الخدمة')
        return redirect('manage_services')

    if request.method == 'POST':
        notes = request.POST.get('approval_notes', '')

        # Use the model's approve method
        service.approve(user, notes)

        # Send email notification to creator
        try:
            creator_profile = service.creator.userprofile.Coach_profile
            subject = f"تم قبول خدمتك: {service.title}"
            context = {
                'service': service,
                'creator': creator_profile,
                'notes': notes,
                'club': club
            }
            html_message = render_to_string('services/emails/service_approved.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[service.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم قبول الخدمة "{service.title}" بنجاح')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم قبول الخدمة بنجاح'})

    return redirect('manage_services')


@login_required
def reject_service(request, service_id):
    """Reject a service"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذه الخدمة')
        return redirect('manage_services')

    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')

        # Use the model's reject method
        service.reject(user, rejection_reason)

        # Send email notification to creator
        try:
            creator_profile = service.creator.userprofile.Coach_profile
            subject = f"تم رفض خدمتك: {service.title}"
            context = {
                'service': service,
                'creator': creator_profile,
                'rejection_reason': rejection_reason,
                'club': club
            }
            html_message = render_to_string('services/emails/service_rejected.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[service.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم رفض الخدمة "{service.title}"')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم رفض الخدمة'})

    return redirect('manage_services')


@login_required
def service_detail(request, service_id):
    """View detailed information about a service"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بعرض هذه الخدمة')
        return redirect('manage_services')

    # Get service coaches

    context = {
        'service': service,
        'creator_profile': service.creator.userprofile.Coach_profile if hasattr(service.creator,
                                                                                'userprofile') else None,
        'club': club,
        'LANGUAGE_CODE': translation.get_language()
    }

    return render(request, 'club_dashboard/services/service_detail.html', context)


@login_required
def bulk_approve_services(request):
    """Bulk approve multiple services"""
    if request.method == 'POST':
        service_ids = request.POST.getlist('service_ids')
        notes = request.POST.get('bulk_notes', '')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(
            user.userprofile.administrator_profile, 'club', None) or getattr(user.userprofile.vendor_manager_profile,
                                                                             'club', None)

        if not service_ids:
            messages.error(request, 'لم يتم تحديد أي خدمات')
            return redirect('manage_services')

        # Get services to update
        services_to_approve = ServicesModel.objects.filter(
            id__in=service_ids,
            club=club,
            approval_status='pending'
        )

        updated_count = 0
        for service in services_to_approve:
            service.approve(user, notes)
            updated_count += 1

        messages.success(request, f'تم قبول {updated_count} خدمة بنجاح')

    return redirect('manage_services')


@login_required
def bulk_reject_services(request):
    """Bulk reject multiple services"""
    if request.method == 'POST':
        service_ids = request.POST.getlist('service_ids')
        rejection_reason = request.POST.get('bulk_rejection_reason', 'تم الرفض بواسطة الإدارة')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(
            user.userprofile.administrator_profile, 'club', None) or getattr(user.userprofile.vendor_manager_profile,
                                                                             'club', None)

        if not service_ids:
            messages.error(request, 'لم يتم تحديد أي خدمات')
            return redirect('manage_services')

        # Get services to update
        services_to_reject = ServicesModel.objects.filter(
            id__in=service_ids,
            club=club,
            approval_status='pending'
        )

        updated_count = 0
        for service in services_to_reject:
            service.reject(user, rejection_reason)
            updated_count += 1

        messages.success(request, f'تم رفض {updated_count} خدمة')

    return redirect('manage_services')


@login_required
def toggle_service_status(request, service_id):
    """Toggle service enabled/disabled status"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذه الخدمة')
        return redirect('manage_services')

    if request.method == 'POST':
        service.is_enabled = not service.is_enabled
        service.save()

        status_text = 'تم تفعيل' if service.is_enabled else 'تم إلغاء تفعيل'
        messages.success(request, f'{status_text} الخدمة "{service.title}"')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'{status_text} الخدمة',
                'is_enabled': service.is_enabled
            })

    return redirect('manage_services')


def delete_service(request, service_id):
    if request.method == 'POST':
        service = get_object_or_404(ServicesModel, id=service_id)
        # Add any permission checks here
        service.delete()
        messages.success(request, 'Service deleted successfully')
        return redirect('manage_services')


# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.urls import reverse
from decimal import Decimal
import json

from .models import RefundDispute, RefundDisputeAttachment, RefundStatus, RefundType, DisputeType
from students.models import Order  # Assuming Order is in students app
from .forms import RefundDisputeForm, RefundDecisionForm, RefundAttachmentForm
from django.utils.translation import get_language


@login_required
@club_permission_required('refund_dashboard')
def refund_dashboard(request):
    """Main dashboard for refund and dispute management"""
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    dispute_type_filter = request.GET.get('dispute_type', '')
    priority_filter = request.GET.get('priority', '')
    search_query = request.GET.get('search', '')

    # Base queryset - exclude PENDING status by default
    disputes = RefundDispute.objects.select_related(
        'deal', 'client', 'vendor', 'reviewed_by'
    ).prefetch_related('attachments').exclude(status=RefundStatus.PENDING)

    # Apply filters - but allow showing pending if explicitly filtered
    if status_filter:
        disputes = RefundDispute.objects.select_related(
            'deal', 'client', 'vendor', 'reviewed_by'
        ).prefetch_related('attachments').filter(status=status_filter)
    else:
        # Default filter (exclude pending)
        disputes = disputes.exclude(status=RefundStatus.PENDING)

    if dispute_type_filter:
        disputes = disputes.filter(dispute_type=dispute_type_filter)

    if priority_filter:
        disputes = disputes.filter(priority=priority_filter)

    if search_query:
        disputes = disputes.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(deal__id__icontains=search_query) |
            Q(client__email__icontains=search_query) |
            Q(vendor__email__icontains=search_query)
        )

    # Get statistics - adjust to exclude pending from totals
    total_disputes = RefundDispute.objects.exclude(status=RefundStatus.PENDING).count()
    pending_disputes = RefundDispute.objects.filter(status=RefundStatus.PENDING).count()

    stats = {
        'total_disputes': total_disputes,
        'pending_disputes': pending_disputes,
        'overdue_disputes': RefundDispute.objects.filter(
            status__in=[RefundStatus.INVESTIGATING],
            created_at__lt=timezone.now() - timezone.timedelta(days=7)
        ).count(),
        'total_refund_amount': RefundDispute.objects.filter(
            status=RefundStatus.APPROVED
        ).aggregate(
            total=Sum('approved_refund_amount')
        )['total'] or Decimal('0.00'),
        'avg_resolution_days': 5.2,
    }

    # Pagination
    paginator = Paginator(disputes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'disputes': page_obj,
        'stats': stats,
        'status_choices': RefundStatus.choices,
        'dispute_type_choices': DisputeType.choices,
        'priority_choices': [('low', 'Low'), ('medium', 'Medium'), ('high', 'High'), ('urgent', 'Urgent')],
        'current_filters': {
            'status': status_filter,
            'dispute_type': dispute_type_filter,
            'priority': priority_filter,
            'search': search_query,
        },
        'LANGUAGE_CODE': get_language(),
    }

    return render(request, 'club_dashboard/refunds/dashboard.html', context)


@login_required
def refund_detail(request, dispute_id):
    """Detailed view of a specific refund/dispute"""

    dispute = get_object_or_404(
        RefundDispute.objects.select_related(
            'deal', 'client', 'vendor', 'reviewed_by'
        ).prefetch_related('attachments'),
        id=dispute_id
    )

    # Get order details if available
    order_items = []
    if dispute.deal:
        order_items = dispute.deal.items.select_related('product', 'service').all()

    context = {
        'dispute': dispute,
        'order_items': order_items,
        'can_approve': dispute.can_be_approved(),
        'can_reject': dispute.can_be_rejected(),
        'can_resolve': dispute.can_be_resolved(),
    }

    return render(request, 'club_dashboard/refunds/detail.html', context)


@login_required
@require_http_methods(["POST"])
def approve_refund(request, dispute_id):
    """Approve a refund dispute"""

    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if not dispute.can_be_approved():
        return JsonResponse({
            'success': False,
            'error': 'This dispute cannot be approved in its current state'
        })

    try:
        # Get form data
        approved_amount = Decimal(request.POST.get('approved_amount', '0'))
        vendor_percentage = Decimal(request.POST.get('vendor_percentage', '0'))
        client_percentage = Decimal(request.POST.get('client_percentage', '100'))
        admin_notes = request.POST.get('admin_notes', '')

        # Validate percentages
        if vendor_percentage + client_percentage != 100:
            return JsonResponse({
                'success': False,
                'error': 'Vendor and client percentages must sum to 100%'
            })

        # Validate approved amount
        if approved_amount > dispute.requested_refund_amount:
            return JsonResponse({
                'success': False,
                'error': 'Approved amount cannot exceed requested amount'
            })

        # Update dispute
        dispute.status = RefundStatus.APPROVED
        dispute.approved_refund_amount = approved_amount
        dispute.vendor_percentage = vendor_percentage
        dispute.client_percentage = client_percentage
        dispute.admin_notes = admin_notes
        dispute.reviewed_by = request.user
        dispute.approved_at = timezone.now()
        dispute.save()

        # Process the actual refund (integrate with payment system)
        process_refund_payment(dispute)

        return redirect('detail', dispute_id=dispute.id)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing approval: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def reject_refund(request, dispute_id):
    """Reject a refund dispute"""

    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if not dispute.can_be_rejected():
        return JsonResponse({
            'success': False,
            'error': 'This dispute cannot be rejected in its current state'
        })

    try:
        rejection_reason = request.POST.get('rejection_reason', '')
        admin_notes = request.POST.get('admin_notes', '')

        if not rejection_reason:
            return JsonResponse({
                'success': False,
                'error': 'Rejection reason is required'
            })

        dispute.status = RefundStatus.REJECTED
        dispute.rejection_reason = rejection_reason
        dispute.admin_notes = admin_notes
        dispute.reviewed_by = request.user
        dispute.rejected_at = timezone.now()
        dispute.save()

        # Send notification to client
        send_refund_rejection_notification(dispute)

        return redirect('detail', dispute_id=dispute.id)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing rejection: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def mark_investigating(request, dispute_id):
    """Mark dispute as under investigation"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if dispute.status != RefundStatus.PENDING:
        return JsonResponse({
            'success': False,
            'error': 'Only pending disputes can be marked as investigating'
        })

    dispute.status = RefundStatus.INVESTIGATING
    dispute.requires_investigation = True
    dispute.reviewed_by = request.user
    dispute.save()

    return redirect('detail', dispute_id=dispute.id)


@login_required
@require_http_methods(["POST"])
def resolve_dispute(request, dispute_id):
    """Mark dispute as resolved"""

    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if not dispute.can_be_resolved():
        return JsonResponse({
            'success': False,
            'error': 'This dispute cannot be resolved in its current state'
        })

    resolution_notes = request.POST.get('resolution_notes', '')

    dispute.status = RefundStatus.RESOLVED
    dispute.resolution_notes = resolution_notes
    dispute.resolved_at = timezone.now()
    dispute.save()

    return redirect('detail', dispute_id=dispute.id)


@login_required
def bulk_action(request):
    """Handle bulk actions on multiple disputes"""

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    try:
        data = json.loads(request.body)
        dispute_ids = data.get('dispute_ids', [])
        action = data.get('action', '')

        if not dispute_ids or not action:
            return JsonResponse({
                'success': False,
                'error': 'Missing dispute IDs or action'
            })

        disputes = RefundDispute.objects.filter(id__in=dispute_ids)

        if action == 'mark_investigating':
            disputes.filter(status=RefundStatus.PENDING).update(
                status=RefundStatus.INVESTIGATING,
                requires_investigation=True,
                reviewed_by=request.user
            )
            message = f'{disputes.count()} disputes marked as investigating'

        elif action == 'assign_priority':
            priority = data.get('priority', 'medium')
            disputes.update(priority=priority)
            message = f'{disputes.count()} disputes priority updated to {priority}'

        else:
            return JsonResponse({
                'success': False,
                'error': 'Invalid action'
            })

        return JsonResponse({
            'success': True,
            'message': message
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing bulk action: {str(e)}'
        })


@login_required
def export_disputes(request):
    """Export disputes to CSV"""

    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="refund_disputes.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Title', 'Status', 'Dispute Type', 'Client', 'Vendor',
        'Original Amount', 'Requested Amount', 'Approved Amount',
        'Created At', 'Resolved At', 'Priority'
    ])

    disputes = RefundDispute.objects.select_related('client', 'vendor').all()

    for dispute in disputes:
        writer.writerow([
            dispute.id,
            dispute.title,
            dispute.get_status_display(),
            dispute.get_dispute_type_display(),
            dispute.client.username if dispute.client else '',
            dispute.vendor.username if dispute.vendor else '',
            dispute.original_amount,
            dispute.requested_refund_amount,
            dispute.approved_refund_amount or '',
            dispute.created_at.strftime('%Y-%m-%d %H:%M'),
            dispute.resolved_at.strftime('%Y-%m-%d %H:%M') if dispute.resolved_at else '',
            dispute.get_priority_display()
        ])

    return response


# Helper functions
def process_refund_payment(dispute):
    """Process the actual refund payment"""
    # This should integrate with your payment processor
    # For now, just a placeholder
    try:
        # Example: integrate with Stripe, PayPal, etc.
        # payment_processor.refund(
        #     transaction_id=dispute.deal.payment_id,
        #     amount=dispute.approved_refund_amount
        # )

        # Log the refund
        print(f"Processing refund of {dispute.approved_refund_amount} for dispute {dispute.id}")

        # Update order status if needed
        if dispute.deal and dispute.is_full_refund():
            dispute.deal.status = 'refunded'  # Add this status to your Order model
            dispute.deal.save()

    except Exception as e:
        print(f"Error processing refund payment: {e}")
        # You might want to log this error and possibly revert the dispute status


def send_refund_rejection_notification(dispute):
    """Send notification when refund is rejected"""
    # Implement email/SMS notification logic
    try:

        from django.core.mail import send_mail
        from django.conf import settings

        subject = f"Refund Request Rejected - Order #{dispute.deal.id}"
        message = f"""
        Dear {dispute.client.get_full_name()},

        Your refund request for Order #{dispute.deal.id} has been rejected.

        Reason: {dispute.rejection_reason}

        If you have questions, please contact our support team.

        Best regards,
        Club Management Team
        """

        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [dispute.client.email],
            fail_silently=True,
        )

    except Exception as e:
        print(f"Error sending rejection notification: {e}")


# API Views for AJAX calls
@login_required
def get_dispute_stats(request):
    """Get dispute statistics for dashboard"""

    stats = {
        'pending': RefundDispute.objects.filter(status=RefundStatus.PENDING).count(),
        'investigating': RefundDispute.objects.filter(status=RefundStatus.INVESTIGATING).count(),
        'approved': RefundDispute.objects.filter(status=RefundStatus.APPROVED).count(),
        'rejected': RefundDispute.objects.filter(status=RefundStatus.REJECTED).count(),
        'resolved': RefundDispute.objects.filter(status=RefundStatus.RESOLVED).count(),
        'overdue': RefundDispute.objects.filter(
            status__in=[RefundStatus.PENDING, RefundStatus.INVESTIGATING],
            created_at__lt=timezone.now() - timezone.timedelta(days=7)
        ).count(),
    }

    return JsonResponse(stats)


@login_required
def update_dispute_priority(request, dispute_id):
    """Update dispute priority"""

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    dispute = get_object_or_404(RefundDispute, id=dispute_id)
    priority = request.POST.get('priority')

    if priority not in ['low', 'medium', 'high', 'urgent']:
        return JsonResponse({'success': False, 'error': 'Invalid priority'})

    dispute.priority = priority
    dispute.save()

    return JsonResponse({
        'success': True,
        'message': f'Priority updated to {priority}'
    })


# Helper function
def get_vendor_from_order(order):
    """Get vendor from order - you'll need to implement this based on your Order model"""
    try:
        for item in order.items.filter(product__isnull=False):
            if (item.product.creator and
                    hasattr(item.product.creator, 'userprofile') and
                    hasattr(item.product.creator.userprofile, 'Coach_profile')):
                return item.product.creator
        return None
    except Exception:
        return None


from django.core.exceptions import PermissionDenied


@login_required
def coach_details(request, coach_id):
    context = {}
    try:
        coach = CoachProfile.objects.get(id=coach_id)
        user_profile = coach.userprofile_set.first()

        # Format working hours for display
        working_hours_display = []
        DAYS = [
            ('monday', 'الإثنين', 'Monday'),
            ('tuesday', 'الثلاثاء', 'Tuesday'),
            ('wednesday', 'الأربعاء', 'Wednesday'),
            ('thursday', 'الخميس', 'Thursday'),
            ('friday', 'الجمعة', 'Friday'),
            ('saturday', 'السبت', 'Saturday'),
            ('sunday', 'الأحد', 'Sunday'),
        ]

        for day_code, day_ar, day_en in DAYS:
            day_data = coach.working_hours.get(day_code, {})
            working_hours_display.append({
                'day_code': day_code,
                'day_name': day_ar if translation.get_language() == 'ar' else day_en,
                'enabled': day_code in coach.working_hours,
                'opening': day_data.get('opening', '--:--'),
                'closing': day_data.get('closing', '--:--'),
            })

        context = {
            'coach': coach,
            'user_profile': user_profile,
            'working_hours': working_hours_display,
            'is_working_hours_enabled': coach.is_working_hours_enabled,
            'approval_statuses': dict(CoachProfile.APPROVAL_STATUS_CHOICES),
            'business_document_types': dict(CoachProfile.BUSINESS_DOCUMENT_CHOICES),
            'club': coach.club,
            'LANGUAGE_CODE': translation.get_language(),
        }

    except CoachProfile.DoesNotExist:
        messages.error(request, "Coach not found")
        return redirect('viewCoachs')

    return render(request, 'club_dashboard/coachs/coach_details.html', context)


from .forms import CustomRoleForm


@login_required
def add_custom_role(request):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile,
                                                                               'club', None) or getattr(
        user.userprofile.vendor_manager_profile, 'club', None)

    language_code = translation.get_language()

    if request.method == 'POST':
        form = CustomRoleForm(request.POST, language_code=language_code)
        if form.is_valid():
            role = form.save(commit=False)
            role.club = club
            role.save()
            return redirect('viewDirectors')
    else:
        form = CustomRoleForm(language_code=language_code)

    context = {
        'form': form,
        'club': club,
        'LANGUAGE_CODE': language_code
    }
    return render(request, 'club_dashboard/directors/add_custom_role.html', context)


# views.py
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import translation
from .models import CustomRole
from .forms import CustomRoleForm
from django.http import HttpResponseForbidden


def view_custom_roles(request):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or \
           getattr(user.userprofile.administrator_profile, 'club', None) or \
           getattr(user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        return HttpResponseForbidden("You don't have permission to view this page")

    custom_roles = CustomRole.objects.filter(club=club, is_active=True).order_by('-created_at')

    # Count users for each role
    role_user_counts = []
    for role in custom_roles:
        count = UserProfile.objects.filter(custom_role=role).count()
        role_user_counts.append({
            'role': role,
            'user_count': count
        })

    context = {
        'role_user_counts': role_user_counts,
        'LANGUAGE_CODE': translation.get_language()
    }
    return render(request, 'club_dashboard/directors/view_custom_roles.html', context)


def edit_custom_role(request, id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or \
           getattr(user.userprofile.administrator_profile, 'club', None) or \
           getattr(user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        return HttpResponseForbidden("You don't have permission to view this page")

    custom_role = get_object_or_404(CustomRole, id=id, club=club)

    language_code = translation.get_language()

    if request.method == 'POST':
        form = CustomRoleForm(request.POST, instance=custom_role, language_code=language_code)
        if form.is_valid():
            form.save()
            messages.success(request, "Role updated successfully!")
            return redirect('view_custom_roles')
    else:
        form = CustomRoleForm(instance=custom_role, language_code=language_code)

    # Get permission choices based on current language
    permission_choices = CustomRole.get_permission_choices(language_code)

    # Create a dictionary of permission statuses
    permission_status = {}
    for perm_code, perm_name in permission_choices:
        permission_status[perm_code] = custom_role.has_permission(perm_code)

    context = {
        'form': form,
        'role': custom_role,
        'permission_status': permission_status,
        'permission_choices': permission_choices,
        'LANGUAGE_CODE': language_code
    }
    return render(request, 'club_dashboard/directors/edit_custom_role.html', context)


@club_permission_required('deleteDirector')
def delete_custom_role(request, id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or \
           getattr(user.userprofile.administrator_profile, 'club', None) or \
           getattr(user.userprofile.vendor_manager_profile, 'club', None)

    if not club:
        return HttpResponseForbidden("You don't have permission to view this page")

    custom_role = get_object_or_404(CustomRole, id=id, club=club)

    # Get user count before deletion for message
    user_count = UserProfile.objects.filter(custom_role=custom_role).count()

    # Soft delete by setting is_active=False
    custom_role.is_active = False
    custom_role.save()

    if user_count > 0:
        messages.success(request, f"Role deleted successfully! {user_count} user(s) have lost this role's permissions.")
    else:
        messages.success(request, "Role deleted successfully!")

    return redirect('view_custom_roles')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required  # Add this import
from django.utils import timezone, translation  # Add translation import
from django.core.paginator import Paginator  # Add this import
from django.db.models import Q
from datetime import timedelta
from coach_dashboard.models import Promotion, PromotionFeature
from students.models import ProductsModel, ServicesModel
from accounts.models import UserProfile


@login_required
def director_promotions(request):
    """View for directors to manage promotion requests"""
    user = request.user
    user_profile = user.userprofile

    # Verify director access
    if not hasattr(user_profile, 'director_profile'):
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    club = user_profile.director_profile.club
    lang = translation.get_language()

    # Get all promotions for this club
    promotions = Promotion.objects.filter(
        Q(product__club=club) | Q(service__club=club)).order_by('-created_at')

    # Filter by status if requested
    status_filter = request.GET.get('status', 'all')
    if status_filter and status_filter != 'all' and status_filter in dict(Promotion.STATUS_CHOICES).keys():
        promotions = promotions.filter(status=status_filter)
    else:
        promotions = promotions.exclude(status='expired')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        promotions = promotions.filter(
            Q(product__title__icontains=search_query) |
            Q(service__title__icontains=search_query) |
            Q(coach__full_name__icontains=search_query) |
            Q(promotion_package__name__icontains=search_query)
        )

    # Calculate statistics
    total_promotions = promotions.count()
    pending_promotions = promotions.filter(status='pending').count()
    active_promotions = promotions.filter(status='active').count()
    rejected_promotions = promotions.filter(status='rejected').count()

    # Pagination
    paginator = Paginator(promotions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'promotions': page_obj,
        'total_promotions': total_promotions,
        'pending_promotions': pending_promotions,
        'active_promotions': active_promotions,
        'rejected_promotions': rejected_promotions,
        'status_filter': status_filter,
        'search_query': search_query,
        'club': club,
        'LANGUAGE_CODE': lang,
    }
    return render(request, 'club_dashboard/promotions/manage_promotions.html', context)


@login_required
def approve_promotion(request, promotion_id):
    """Approve a promotion request"""
    user = request.user
    user_profile = user.userprofile

    # Verify director access
    if not hasattr(user_profile, 'director_profile'):
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    club = user_profile.director_profile.club
    promotion = get_object_or_404(Promotion, id=promotion_id)

    # Verify promotion belongs to this club
    if (promotion.product and promotion.product.club != club) or (promotion.service and promotion.service.club != club):
        messages.error(request, "This promotion doesn't belong to your club.")
        return redirect('director_promotions')

    if request.method == 'POST':
        approval_notes = request.POST.get('approval_notes', '')
        start_date_str = request.POST.get('start_date')

        try:
            promotion.status = 'active'

            # Set start date from form or now if not provided
            if start_date_str:
                promotion.start_date = timezone.make_aware(datetime.strptime(start_date_str, '%Y-%m-%d'))
            else:
                promotion.start_date = timezone.now()

            # Set end date based on duration
            if promotion.duration_days > 0:
                promotion.end_date = promotion.start_date + timedelta(days=promotion.duration_days)

            if approval_notes:
                promotion.notes = f"Approved: {approval_notes}"

            promotion.save()

            # Send notification to coach
            send_notification(
                user=promotion.coach.user,
                title=f"Promotion Approved",
                message=f"Your promotion for {promotion.get_promotion_item_name()} has been approved.",
                notification_type='promotion_approved'
            )

            messages.success(request, f"Promotion for {promotion.get_promotion_item_name()} has been approved!")

        except Exception as e:
            messages.error(request, f"Error approving promotion: {str(e)}")

    return redirect('director_promotions')


@login_required
def reject_promotion(request, promotion_id):
    """Reject a promotion request"""
    user = request.user
    user_profile = user.userprofile

    # Verify director access
    if not hasattr(user_profile, 'director_profile'):
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    club = user_profile.director_profile.club
    promotion = get_object_or_404(Promotion, id=promotion_id)

    # Verify promotion belongs to this club
    if (promotion.product and promotion.product.club != club) or (promotion.service and promotion.service.club != club):
        messages.error(request, "This promotion doesn't belong to your club.")
        return redirect('director_promotions')

    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        promotion.status = 'rejected'
        promotion.notes = f"Rejected: {rejection_reason}"
        promotion.save()

        # Send notification to coach
        send_notification(
            user=promotion.coach.user,
            title=f"Promotion Rejected",
            message=f"Your promotion for {promotion.get_promotion_item_name()} was rejected. Reason: {rejection_reason}",
            notification_type='promotion_rejected'
        )

        messages.success(request, f"Promotion for {promotion.get_promotion_item_name()} has been rejected.")

    return redirect('director_promotions')


@login_required
def view_pending_promotion(request, promotion_id):
    """View details of a pending promotion"""
    user = request.user
    user_profile = user.userprofile

    # Verify director access
    if not hasattr(user_profile, 'director_profile'):
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    club = user_profile.director_profile.club
    promotion = get_object_or_404(Promotion, id=promotion_id)

    # Verify promotion belongs to this club
    if (promotion.product and promotion.product.club != club) or (promotion.service and promotion.service.club != club):
        messages.error(request, "This promotion doesn't belong to your club.")
        return redirect('director_promotions')

    # Calculate remaining days if active
    remaining_days = promotion.get_remaining_days() if promotion.status == 'active' else 0

    context = {
        'promotion': promotion,
        'club': club,
        'remaining_days': remaining_days,
        'LANGUAGE_CODE': translation.get_language(),
    }
    return render(request, 'club_dashboard/promotions/view_promotion.html', context)


@login_required
def manage_promotion_features(request):
    """Manage promotion features and base price"""
    user = request.user
    user_profile = user.userprofile

    if not hasattr(user_profile, 'director_profile'):
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    club = user_profile.director_profile.club
    features = PromotionFeature.objects.filter(club=club).order_by('name')

    context = {
        'features': features,
        'club': club,
    }
    return render(request, 'club_dashboard/promotions/manage_features.html', context)


@login_required
def save_promotion_feature(request):
    """Create or update a promotion feature"""
    user = request.user
    user_profile = user.userprofile

    if not hasattr(user_profile, 'director_profile'):
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    club = user_profile.director_profile.club

    if request.method == 'POST':
        feature_id = request.POST.get('feature_id')
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        price_multiplier = float(request.POST.get('price_multiplier', 1.0))
        is_active = request.POST.get('is_active') == 'on'

        try:
            if feature_id:  # Update existing feature
                feature = get_object_or_404(PromotionFeature, id=feature_id, club=club)
                feature.name = name
                feature.description = description
                feature.price_multiplier = price_multiplier
                feature.is_active = is_active
                feature.save()
                messages.success(request, f"Feature '{name}' updated successfully!")
            else:  # Create new feature
                PromotionFeature.objects.create(
                    club=club,
                    name=name,
                    description=description,
                    price_multiplier=price_multiplier,
                    is_active=is_active
                )
                messages.success(request, f"Feature '{name}' created successfully!")
        except Exception as e:
            messages.error(request, f"Error saving feature: {str(e)}")

    return redirect('manage_promotion_features')


@login_required
def delete_promotion_feature(request, feature_id):
    """Delete a promotion feature"""
    user = request.user
    user_profile = user.userprofile

    if not hasattr(user_profile, 'director_profile'):
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    club = user_profile.director_profile.club
    feature = get_object_or_404(PromotionFeature, id=feature_id, club=club)

    if request.method == 'POST':
        feature_name = feature.name
        feature.delete()
        messages.success(request, f"Feature '{feature_name}' deleted successfully!")

    return redirect('manage_promotion_features')


from decimal import InvalidOperation


@login_required
def set_promotion_base_price(request):
    """Set the base price per day for promotions"""
    if not hasattr(request.user.userprofile, 'director_profile'):
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    club = request.user.userprofile.director_profile.club

    if request.method == 'POST':
        base_price = request.POST.get('base_price')

        try:
            # Convert to Decimal for proper monetary handling
            base_price = Decimal(base_price)
            if base_price <= 0:
                raise ValueError("Price must be positive")

            club.promotion_base_price = base_price
            club.save()
            messages.success(request, f"Base price set to {base_price} {club.vat_settings.currency} per day")
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"Invalid price value: {str(e)}")

    return redirect('manage_promotion_features')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from students.models import Order
from accountant_dashboard.models import BillRevision, BillRevisionComment
from accountant_dashboard.forms import BillRevisionForm, BillRevisionCommentForm
from django.utils import translation


@login_required
def director_bills_review(request):
    if not hasattr(request.user.userprofile, 'director_profile'):
        return redirect('club_dashboard_index')

    director_profile = request.user.userprofile.director_profile
    club = director_profile.club

    pending_revisions = BillRevision.objects.filter(
        order__club=club,
        status='accountant_reviewed'
    ).select_related('order', 'accountant', 'order__user')

    reviewed_revisions = BillRevision.objects.filter(
        order__club=club,
        status__in=['director_reviewed', 'approved', 'rejected']
    ).select_related('order', 'accountant', 'director', 'order__user').order_by('-updated_at')

    total_revisions = reviewed_revisions.count() + pending_revisions.count()

    context = {
        'club': club,
        'pending_revisions': pending_revisions,
        'reviewed_revisions': reviewed_revisions,
        'total_revisions': total_revisions,
        'LANGUAGE_CODE': translation.get_language()
    }
    return render(request, 'club_dashboard/bills/bills_review.html', context)


from django.utils import translation  # Make sure you imported this

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import translation
from students.models import Order, OrderItem
from accountant_dashboard.models import BillRevision, BillRevisionComment
from accountant_dashboard.forms import BillRevisionForm, BillRevisionCommentForm
from collections import defaultdict


@login_required
def director_review_bill(request, revision_id):
    if not hasattr(request.user.userprofile, 'director_profile'):
        messages.error(request, "You don't have permission to access this page")
        return redirect('club_dashboard_index')

    director_club = request.user.userprofile.director_profile.club
    revision = get_object_or_404(BillRevision, id=revision_id, order__club=director_club)
    order = revision.order

    form = BillRevisionForm(instance=revision)
    comment_form = BillRevisionCommentForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        is_comment_submission = 'comment' in request.POST

        if is_comment_submission:
            comment_form = BillRevisionCommentForm(request.POST)
            if comment_form.is_valid():
                comment_text = comment_form.cleaned_data.get('comment')
                if comment_text:
                    BillRevisionComment.objects.create(
                        revision=revision,
                        author=request.user.userprofile,
                        comment=comment_text
                    )
                    messages.success(request, "Comment added successfully.")
                    return redirect('director_review_bill', revision_id=revision.id)

        elif revision.status == 'accountant_reviewed':
            form = BillRevisionForm(request.POST, instance=revision)
            if form.is_valid():
                updated_revision = form.save(commit=False)
                updated_revision.director = request.user.userprofile

                if action in ['approved', 'rejected']:
                    updated_revision.status = action
                    updated_revision.save()
                    messages.success(request, f"Bill has been {action}.")
                    return redirect('director_bills_review')
                else:
                    messages.error(request, "Invalid action performed.")
        else:
            messages.warning(request, "This bill has already been processed and cannot be modified.")
            return redirect('director_review_bill', revision_id=revision.id)

    # --- Group items by coach ---
    coaches_items = defaultdict(lambda: {'coach_profile': None, 'items': []})
    order_items = order.items.select_related(
        'product__creator__userprofile__Coach_profile',
        'service__creator__userprofile__Coach_profile'
    ).all()

    for item in order_items:
        creator = None
        if item.product and item.product.creator:
            creator = item.product.creator
        elif item.service and item.service.creator:
            creator = item.service.creator

        if creator and hasattr(creator, 'userprofile') and hasattr(creator.userprofile, 'Coach_profile'):
            coach_profile = creator.userprofile.Coach_profile
            coaches_items[coach_profile.id]['coach_profile'] = coach_profile
            coaches_items[coach_profile.id]['items'].append(item)

    comments = revision.comments.all().select_related('author__user')
    context = {
        'club': director_club,
        'order': order,
        'revision': revision,
        'form': form,
        'comment_form': comment_form,
        'comments': comments,
        'coaches_items': dict(coaches_items),  # Pass the grouped items to the template
        'LANGUAGE_CODE': translation.get_language(),
        'can_take_action': revision.status == 'accountant_reviewed'
    }

    return render(request, 'club_dashboard/bills/review_bill.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from .models import VATSettings
from .forms import VATSettingsForm
from .decorators import club_permission_required


@club_permission_required('manage_vat_settings')
@login_required
def manage_vat_settings(request):
    """View for managing VAT settings."""
    club = getattr(request.user.userprofile.director_profile, 'club', None) or \
           getattr(request.user.userprofile.administrator_profile, 'club', None) or \
           getattr(request.user.userprofile.vendor_manager_profile, 'club', None) or \
           getattr(request.user.userprofile.custom_role_profile, 'club', None)

    if not club:
        messages.error(request, 'No club associated with your account.')
        return redirect('club_dashboard_index')

    vat_settings, created = VATSettings.objects.get_or_create(
        club=club,
        defaults={'is_enabled': False, 'percentage': 15.00, 'currency': 'SAR'}
    )

    if request.method == 'POST':
        form = VATSettingsForm(request.POST, instance=vat_settings)
        if form.is_valid():
            form.save()
            if request.LANGUAGE_CODE == 'ar':
                messages.success(request, 'تم حفظ إعدادات ضريبة القيمة المضافة بنجاح')
            else:
                messages.success(request, 'VAT settings saved successfully')
            return redirect('manage_vat_settings')
    else:
        form = VATSettingsForm(instance=vat_settings)

    context = {
        'form': form,
        'vat_settings': vat_settings,
        'club': club,
        'LANGUAGE_CODE': request.LANGUAGE_CODE
    }

    return render(request, 'club_dashboard/director/manage_vat_settings.html', context)


def vendor_edit_application(request, vendor_id):
    """Allow vendor to edit their application after rejection feedback"""

    print("=== vendor_edit_application called ===")
    print("vendor_id:", vendor_id)
    print("request.method:", request.method)

    try:
        print("Trying to fetch vendor with pending status...")
        vendor = get_object_or_404(
            CoachProfile,
            id=vendor_id,
            approval_status='pending'
        )
        print("Vendor found:", vendor)

        if request.method == 'POST':
            print("POST request received")
            from accounts.forms import VendorRegistrationForm

            print("POST data:", request.POST)
            print("FILES data:", request.FILES)

            form = VendorRegistrationForm(
                request.POST,
                request.FILES,
                instance=vendor
            )

            print("Form created, validating...")
            if form.is_valid():
                print("Form is valid ✅")

                vendor = form.save(commit=False)
                print("Vendor instance updated (not saved yet)")

                print("Old approval_notes:", vendor.approval_notes)
                vendor.approval_notes = ""
                print("approval_notes cleared")

                vendor.save()
                print("Vendor saved successfully")

                print("Sending resubmission notification...")
                send_resubmission_notification(vendor)
                print("Notification sent")

                lang = translation.get_language()
                print("Current language:", lang)

                messages.success(
                    request,
                    'تم إعادة تقديم طلبك بنجاح. سيتم مراجعته قريباً.'
                    if lang == 'ar'
                    else 'Your application has been resubmitted successfully. It will be reviewed soon.'
                )

                print("Rendering success template")
                return render(
                    request,
                    'accounts/vendor_resubmission_success.html',
                    {'vendor': vendor}
                )
            else:
                print("❌ Form is NOT valid")
                print("Form errors:", form.errors)

        else:
            print("GET request received")
            from accounts.forms import VendorRegistrationForm
            form = VendorRegistrationForm(instance=vendor)
            print("Form initialized with vendor instance")

        context = {
            'form': form,
            'vendor': vendor,
            'is_resubmission': True,
            'rejection_notes': vendor.approval_notes,
            'LANGUAGE_CODE': translation.get_language(),
        }

        print("Rendering edit application template")
        return render(
            request,
            'accounts/vendor_edit_application.html',
            context
        )

    except CoachProfile.DoesNotExist:
        print("❌ CoachProfile.DoesNotExist exception")
        messages.error(request, 'Vendor application not found or already processed.')
        return redirect('home')

    except Exception as e:
        print("🔥 Unexpected error occurred:", str(e))
        raise


def send_resubmission_notification(vendor):
    """Send notification to director about vendor resubmission"""
    try:
        from django.core.mail import send_mail
        from django.conf import settings

        # Get director email
        director = vendor.club.directorprofile_set.first()
        if director and director.user.email:
            subject = f'إعادة تقديم طلب بائع - {vendor.full_name}' if translation.get_language() == 'ar' else f'Vendor Application Resubmitted - {vendor.full_name}'

            if translation.get_language() == 'ar':
                message = f"""
                تم إعادة تقديم طلب البائع {vendor.full_name} بعد التعديلات المطلوبة.

                يرجى مراجعة الطلب المحدث في لوحة التحكم.
                """
            else:
                message = f"""
                Vendor {vendor.full_name} has resubmitted their application after making the requested changes.

                Please review the updated application in your dashboard.
                """

            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [director.user.email],
                fail_silently=True,
            )
    except Exception as e:
        print(f"Error sending resubmission notification: {e}")